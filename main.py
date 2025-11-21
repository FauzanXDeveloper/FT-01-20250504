import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
import pandas as pd
import pdfplumber
from pathlib import Path
from datetime import datetime
import re
import fitz  # PyMuPDF
from PIL import Image, ImageTk, ImageFilter
import io
import threading
import time
import numpy as np

class PDFtoExcelApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Configure window
        self.title("PDF to Excel Converter - IRISS Report")
        self.geometry("1400x900")
        
        # Set theme
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        
        # Variables
        self.pdf_path = None
        self.pdf_paths = []  # For batch processing
        self.extracted_sections = {}
        self.file_previews = {}  # Store previews per file: {filename: DataFrame}
        self.pdf_document = None
        self.current_page = 0
        self.pdf_x_offset = 0  # X offset for centered PDF
        self.pdf_y_offset = 0  # Y offset for PDF position
        
        # Selection rectangle variables
        self.selection_start = None
        self.selection_rect = None
        self.selecting = False
        self.saved_selections = {}  # Store selections per page: {page_num: [(x1, y1, x2, y2), ...]}
        self.selection_rectangles = []  # Visual rectangles on canvas
        self.temp_selections = {}  # Temporary selections before saving: {page_num: [(bbox, rect_id), ...]}
        self.selection_history = []  # History for undo: [(page_num, bbox, rect_id), ...]
        self.file_risk_grades = {}  # Store risk grades per file: {filename: risk_grade}
        
        # Database variables
        self.database_df = pd.DataFrame()  # Combined database DataFrame
        
        # Loading screen variables
        self.loading_frame = None
        self.loading_gif_frames = []
        self.loading_gif_label = None
        self.loading_progress_label = None
        self.loading_current_frame = 0
        self.loading_animation_job = None
        self.blur_overlay = None
        
        # Create GUI
        self.create_widgets()
    
    def create_widgets(self):
        # Title Label
        title_label = ctk.CTkLabel(
            self,
            text="PDF to Excel Converter - IRISS Credit Report",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=20)
        
        # Button Frame
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=10, padx=20, fill="x")
        
        # Select PDF Button
        self.select_btn = ctk.CTkButton(
            button_frame,
            text="📁 Select PDF File",
            command=self.select_pdf,
            width=150,
            height=40,
            font=ctk.CTkFont(size=14)
        )
        self.select_btn.pack(side="left", padx=10, pady=10)
        
        # Select Multiple PDFs Button
        self.select_multiple_btn = ctk.CTkButton(
            button_frame,
            text="📁 Select Multiple PDFs",
            command=self.select_multiple_pdfs,
            width=150,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color="#1f538d",
            hover_color="#14375e"
        )
        self.select_multiple_btn.pack(side="left", padx=10, pady=10)
        
        # Select Folder Button
        self.select_folder_btn = ctk.CTkButton(
            button_frame,
            text="📂 Select Folder",
            command=self.select_folder,
            width=150,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color="#1f538d",
            hover_color="#14375e"
        )
        self.select_folder_btn.pack(side="left", padx=10, pady=10)
        
        # File count label
        self.file_count_label = ctk.CTkLabel(
            button_frame,
            text="0 files",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        self.file_count_label.pack(side="left", padx=20, pady=10)
        
        # Export Button
        self.export_btn = ctk.CTkButton(
            button_frame,
            text="📊 Export to Excel",
            command=self.export_to_excel,
            width=150,
            height=40,
            font=ctk.CTkFont(size=14),
            state="disabled",
            fg_color="green",
            hover_color="darkgreen"
        )
        self.export_btn.pack(side="left", padx=10, pady=10)
        
        # Batch Process Button
        self.batch_btn = ctk.CTkButton(
            button_frame,
            text="⚡ Batch Process All",
            command=self.batch_process_pdfs,
            width=150,
            height=40,
            font=ctk.CTkFont(size=14),
            state="disabled",
            fg_color="#9333ea",
            hover_color="#7e22ce"
        )
        self.batch_btn.pack(side="left", padx=10, pady=10)
        
        # Create Tabbed Interfaceabel to prevent errors (but don't pack it)
        self.file_label = ctk.CTkLabel(self, text="")
        
        # Create Tabbed Interface
        self.tab_view = ctk.CTkTabview(self, width=1350, height=650)
        self.tab_view.pack(pady=10, padx=20, fill="both", expand=True)
        
        # Tab 1: PDF Viewer
        self.tab_view.add("📄 PDF Viewer")
        pdf_tab = self.tab_view.tab("📄 PDF Viewer")
        
        # PDF navigation frame
        pdf_nav_frame = ctk.CTkFrame(pdf_tab)
        pdf_nav_frame.pack(pady=5, padx=10, fill="x")
        
        # Create a frame for centered buttonsent errors (but don't pack it)
        self.page_label = ctk.CTkLabel(pdf_nav_frame, text="")
        
        # Create a frame for centered buttons
        button_container = ctk.CTkFrame(pdf_nav_frame)
        button_container.pack(pady=5)
        
        # Selection mode button
        self.select_table_btn = ctk.CTkButton(
            button_container,
            text="📐 Select Table Area",
            command=self.toggle_selection_mode,
            width=150,
            state="disabled",
            fg_color="purple",
            hover_color="darkviolet"
        )
        self.select_table_btn.pack(side="left", padx=5)
        
        # Save Selection button
        self.save_selection_btn = ctk.CTkButton(
            button_container,
            text="💾 Save Selections",
            command=self.save_current_page_selections,
            width=150,
            state="disabled",
            fg_color="#16a34a",
            hover_color="#15803d"
        )
        self.save_selection_btn.pack(side="left", padx=5)
        
        # Undo Selection button (was Clear Selections)
        self.clear_selections_btn = ctk.CTkButton(
            button_container,
            text="↶ Undo Selection",
            command=self.undo_last_selection,
            width=150,
            state="disabled",
            fg_color="#dc2626",
            hover_color="#991b1b"
        )
        self.clear_selections_btn.pack(side="left", padx=5)
        
        # Auto Extract All Tables button
        self.auto_extract_btn = ctk.CTkButton(
            button_container,
            text="⚡ Auto Extract All Tables",
            command=self.auto_extract_all_tables,
            width=180,
            state="disabled",
            fg_color="#ea580c",
            hover_color="#c2410c"
        )
        self.auto_extract_btn.pack(side="left", padx=5)
        
        # Clear Preview button
        self.clear_preview_btn = ctk.CTkButton(
            button_container,
            text="🧹 Clear Preview",
            command=self.clear_all_previews,
            width=150,
            state="disabled",
            fg_color="#64748b",
            hover_color="#475569"
        )
        self.clear_preview_btn.pack(side="left", padx=5)
        
        # Apply to All Files button
        self.apply_all_btn = ctk.CTkButton(
            button_container,
            text="📋 Apply to All Files",
            command=self.apply_selections_to_all_files,
            width=150,
            state="disabled",
            fg_color="#0891b2",
            hover_color="#0e7490"
        )
        self.apply_all_btn.pack(side="left", padx=5)
        
        # PDF display canvas with scrollbar
        pdf_canvas_frame = ctk.CTkFrame(pdf_tab)
        pdf_canvas_frame.pack(pady=5, padx=10, fill="both", expand=True)
        
        # Create canvas for PDF display
        self.pdf_canvas = ctk.CTkCanvas(
            pdf_canvas_frame,
            bg="#2b2b2b",
            highlightthickness=0
        )
        self.pdf_scrollbar = ctk.CTkScrollbar(
            pdf_canvas_frame,
            command=self.pdf_canvas.yview
        )
        self.pdf_canvas.configure(yscrollcommand=self.pdf_scrollbar.set)
        
        self.pdf_scrollbar.pack(side="right", fill="y")
        self.pdf_canvas.pack(side="left", fill="both", expand=True)

        # Sticky nav buttons overlay (appear over canvas)
        # Placed in the same frame so they float above the canvas content
        self.sticky_prev = ctk.CTkButton(pdf_canvas_frame, text="◀ Previous", width=80, command=self.previous_page, state="disabled")
        self.sticky_prev.place(relx=0.02, rely=0.98, anchor="sw")
        
        # Center sticky button for page info
        self.sticky_center = ctk.CTkLabel(pdf_canvas_frame, text="No PDF loaded", width=120, 
                                         font=ctk.CTkFont(size=11), 
                                         fg_color="#2b2b2b", corner_radius=6)
        self.sticky_center.place(relx=0.5, rely=0.98, anchor="s")
        
        self.sticky_next = ctk.CTkButton(pdf_canvas_frame, text="Next ▶", width=80, command=self.next_page, state="disabled")
        self.sticky_next.place(relx=0.98, rely=0.98, anchor="se")
        
        # Bind mouse events for table selection
        self.pdf_canvas.bind("<ButtonPress-1>", self.on_mouse_down)
        self.pdf_canvas.bind("<B1-Motion>", self.on_mouse_drag)
        self.pdf_canvas.bind("<ButtonRelease-1>", self.on_mouse_up)
        
        # Bind mouse wheel for scrolling
        self.pdf_canvas.bind("<MouseWheel>", self.on_mouse_wheel)
        
        # Tab 2: Excel Preview
        self.tab_view.add("📊 Excel Preview")
        excel_tab = self.tab_view.tab("📊 Excel Preview")
        
        # Excel sheet selector
        excel_nav_frame = ctk.CTkFrame(excel_tab)
        excel_nav_frame.pack(pady=5, padx=10, fill="x")
        
        ctk.CTkLabel(
            excel_nav_frame,
            text="Select File:",
            font=ctk.CTkFont(size=12, weight="bold")
        ).pack(side="left", padx=5)
        
        self.sheet_selector = ctk.CTkComboBox(
            excel_nav_frame,
            values=["No data available"],
            command=self.on_file_selected,
            width=400,
            state="disabled"
        )
        self.sheet_selector.pack(side="left", padx=10)
        
        # Excel data display with scrollbar
        excel_display_frame = ctk.CTkFrame(excel_tab)
        excel_display_frame.pack(pady=5, padx=10, fill="both", expand=True)
        
        # Create treeview with scrollbars for Excel-like display
        tree_scroll_y = ctk.CTkScrollbar(excel_display_frame, orientation="vertical")
        tree_scroll_y.pack(side="right", fill="y")
        
        tree_scroll_x = ctk.CTkScrollbar(excel_display_frame, orientation="horizontal")
        tree_scroll_x.pack(side="bottom", fill="x")
        
        # Use tkinter Treeview for table display
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",
                        background="#2b2b2b",
                        foreground="white",
                        fieldbackground="#2b2b2b",
                        borderwidth=1,
                        relief="solid")
        style.configure("Treeview.Heading",
                        background="#1f538d",
                        foreground="white",
                        borderwidth=1)
        style.map("Treeview",
                  background=[("selected", "#144870")])
        
        self.excel_display = ttk.Treeview(
            excel_display_frame,
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            show="headings",
            selectmode="extended"
        )
        self.excel_display.pack(pady=10, padx=10, fill="both", expand=True)
        
        tree_scroll_y.configure(command=self.excel_display.yview)
        tree_scroll_x.configure(command=self.excel_display.xview)
        
        # Tab 3: Database Viewer
        self.tab_view.add("🗄️ Database Viewer")
        db_tab = self.tab_view.tab("🗄️ Database Viewer")
        
        # Database navigation frame
        db_nav_frame = ctk.CTkFrame(db_tab)
        db_nav_frame.pack(pady=5, padx=10, fill="x")
        
        ctk.CTkLabel(
            db_nav_frame,
            text="Database Builder - All extracted files will be combined",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(side="left", padx=10)
        
        # Button frame for database actions
        db_btn_frame = ctk.CTkFrame(db_nav_frame)
        db_btn_frame.pack(side="right", padx=10)
        
        self.db_build_btn = ctk.CTkButton(
            db_btn_frame,
            text="🔨 Build Database",
            command=self.build_database,
            width=140,
            state="disabled",
            fg_color="#9333ea",
            hover_color="#7e22ce"
        )
        self.db_build_btn.pack(pady=4)
        
        self.db_export_btn = ctk.CTkButton(
            db_btn_frame,
            text="💾 Export Database",
            command=self.export_database,
            width=140,
            state="disabled",
            fg_color="green",
            hover_color="darkgreen"
        )
        self.db_export_btn.pack(pady=4)
        

        
        # Database preview area (treeview)
        db_preview_frame = ctk.CTkFrame(db_tab)
        db_preview_frame.pack(pady=5, padx=10, fill="both", expand=True)
        
        db_scroll_y = ctk.CTkScrollbar(db_preview_frame, orientation="vertical")
        db_scroll_y.pack(side="right", fill="y")
        
        db_scroll_x = ctk.CTkScrollbar(db_preview_frame, orientation="horizontal")
        db_scroll_x.pack(side="bottom", fill="x")
        
        self.db_tree = ttk.Treeview(
            db_preview_frame,
            yscrollcommand=db_scroll_y.set,
            xscrollcommand=db_scroll_x.set,
            show="headings",
            selectmode="extended"
        )
        self.db_tree.pack(pady=10, padx=10, fill="both", expand=True)
        
        db_scroll_y.configure(command=self.db_tree.yview)
        db_scroll_x.configure(command=self.db_tree.xview)
        
        # Status bar
        self.status_label = ctk.CTkLabel(
            self,
            text="Ready - Please select a PDF file to begin",
            font=ctk.CTkFont(size=11),
            anchor="w"
        )
        self.status_label.pack(side="bottom", fill="x", padx=20, pady=10)
        
        # Setup loading screen
        self.setup_loading_screen()
    
    def extract_risk_grade_from_score(self, score):
        """Convert i-SCORE to Risk Grade based on the official ranges"""
        if score < 360:
            return 1
        elif 361 <= score <= 420:
            return 2
        elif 421 <= score <= 460:
            return 3
        elif 461 <= score <= 540:
            return 4
        elif 541 <= score <= 580:
            return 5
        elif 581 <= score <= 620:
            return 6
        elif 621 <= score <= 660:
            return 7
        elif 661 <= score <= 700:
            return 8
        elif 701 <= score <= 740:
            return 9
        elif score > 741:
            return 10
        else:
            return None
    
    def extract_iscore_and_risk_grade(self, pdf_path):
        """Extract i-SCORE and calculate risk grade from PDF"""
        try:
            doc = fitz.open(pdf_path)
            
            for page_num, page in enumerate(doc):
                # Extract text to find i-SCORE
                text_blocks = page.get_text("dict")
                
                for block in text_blocks["blocks"]:
                    if "lines" in block:
                        for line in block["lines"]:
                            if "spans" in line:
                                for span in line["spans"]:
                                    text = span["text"].strip()
                                    
                                    # Look for i-SCORE pattern
                                    if "i-SCORE" in text or "i-score" in text.lower():
                                        # Get the bbox area around i-SCORE
                                        bbox = span["bbox"]
                                        search_rect = fitz.Rect(
                                            bbox[0] - 50, bbox[1] - 20,
                                            bbox[0] + 200, bbox[1] + 50
                                        )
                                        
                                        # Extract text in this region
                                        region_text = page.get_textbox(search_rect)
                                        
                                        # Find the score number
                                        score_match = re.search(r'i-SCORE[^\d]*(\d{3,4})', region_text, re.IGNORECASE)
                                        if score_match:
                                            score = int(score_match.group(1))
                                            risk_grade = self.extract_risk_grade_from_score(score)
                                            doc.close()
                                            return score, risk_grade
                                        
                                        # Alternative pattern - look for numbers near i-SCORE
                                        numbers = re.findall(r'\b(\d{3,4})\b', region_text)
                                        for num_str in numbers:
                                            num = int(num_str)
                                            if 300 <= num <= 900:  # Valid score range
                                                risk_grade = self.extract_risk_grade_from_score(num)
                                                doc.close()
                                                return num, risk_grade
            
            doc.close()
            return None, None
            
        except Exception as e:
            print(f"Error extracting risk grade from {pdf_path}: {e}")
            return None, None
    
    def setup_loading_screen(self):
        """Setup the loading screen with GIF animation and blur effect"""
        try:
            # Load GIF frames
            gif_path = r"C:\Users\aaafauzan\Desktop\VS Code\Pdf to Excel\Picture\Loading.gif"
            gif_image = Image.open(gif_path)
            
            # Extract all frames from GIF
            self.loading_gif_frames = []
            try:
                while True:
                    # Resize frame to reasonable size
                    frame = gif_image.copy().resize((100, 100), Image.Resampling.LANCZOS)
                    # Convert to CTkImage to avoid warnings
                    ctk_frame = ctk.CTkImage(frame, size=(100, 100))
                    self.loading_gif_frames.append(ctk_frame)
                    gif_image.seek(len(self.loading_gif_frames))
            except EOFError:
                pass  # End of GIF frames
                
        except Exception as e:
            print(f"Warning: Could not load loading GIF: {e}")
            # Create a simple default loading image
            default_img = Image.new('RGBA', (100, 100), (100, 100, 100, 255))
            self.loading_gif_frames = [ctk.CTkImage(default_img, size=(100, 100))]
    
    def show_loading_screen(self, message="Loading..."):
        """Show loading screen with blur effect"""
        if self.loading_frame:
            return  # Already showing
            
        # Create blur overlay
        self.blur_overlay = ctk.CTkFrame(
            self,
            fg_color=("gray90", "gray10"),
            bg_color="transparent"
        )
        self.blur_overlay.place(x=0, y=0, relwidth=1, relheight=1)
        
        # Create loading frame (centered)
        self.loading_frame = ctk.CTkFrame(
            self.blur_overlay,
            width=300,
            height=200,
            corner_radius=20,
            fg_color=("white", "#2b2b2b"),
            border_width=2,
            border_color=("gray70", "gray30")
        )
        self.loading_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        # Loading message
        self.loading_message_label = ctk.CTkLabel(
            self.loading_frame,
            text=message,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=("black", "white")
        )
        self.loading_message_label.pack(pady=20)
        
        # GIF animation label
        self.loading_gif_label = ctk.CTkLabel(
            self.loading_frame,
            text="",
            width=100,
            height=100
        )
        self.loading_gif_label.pack(pady=10)
        
        # Progress label
        self.loading_progress_label = ctk.CTkLabel(
            self.loading_frame,
            text="0%",
            font=ctk.CTkFont(size=14),
            text_color=("black", "white")
        )
        self.loading_progress_label.pack(pady=10)
        
        # Start GIF animation
        self.loading_current_frame = 0
        self.animate_loading_gif()
        
        # Update display (use update_idletasks to prevent blocking)
        self.update_idletasks()
    
    def animate_loading_gif(self):
        """Animate the loading GIF"""
        if not self.loading_gif_label or not self.loading_gif_frames:
            return
        
        try:
            # Update frame
            if self.loading_current_frame >= len(self.loading_gif_frames):
                self.loading_current_frame = 0
                
            self.loading_gif_label.configure(image=self.loading_gif_frames[self.loading_current_frame])
            self.loading_current_frame += 1
            
            # Schedule next frame (50ms = 20fps for smoother animation)
            self.loading_animation_job = self.after(50, self.animate_loading_gif)
        except Exception:
            # Silently handle errors if label is destroyed
            pass
    
    def update_loading_progress(self, progress, message=None):
        """Update loading progress (0-100)"""
        try:
            if self.loading_progress_label:
                self.loading_progress_label.configure(text=f"{progress}%")
                
            if message and hasattr(self, 'loading_message_label'):
                self.loading_message_label.configure(text=message)
                
            # Use update_idletasks() to prevent blocking
            self.update_idletasks()
        except Exception:
            # Silently handle errors
            pass
    
    def hide_loading_screen(self):
        """Hide loading screen"""
        # Stop animation
        if self.loading_animation_job:
            self.after_cancel(self.loading_animation_job)
            self.loading_animation_job = None
            
        # Destroy loading elements
        if self.loading_frame:
            self.loading_frame.destroy()
            self.loading_frame = None
            
        if self.blur_overlay:
            self.blur_overlay.destroy()
            self.blur_overlay = None
            
        self.loading_gif_label = None
        self.loading_progress_label = None
    
    def select_pdf(self):
        """Open file dialog to select PDF file"""
        file_path = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        
        if file_path:
            self.pdf_path = file_path
            self.pdf_paths = []  # Clear batch list
            self.file_label.configure(text=Path(file_path).name)
            self.file_count_label.configure(text="1 file")
            self.batch_btn.configure(state="disabled")
            self.apply_all_btn.configure(state="disabled")
            self.status_label.configure(text=f"Selected: {Path(file_path).name}")
            self.export_btn.configure(state="disabled")
            
            # Clear previous extractions
            self.extracted_sections = {}
            
            # Load PDF for viewing
            self.load_pdf_viewer(file_path)
    
    def select_multiple_pdfs(self):
        """Open file dialog to select multiple PDF files"""
        file_paths = filedialog.askopenfilenames(
            title="Select Multiple PDF Files",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        
        if file_paths:
            self.show_loading_screen("Loading PDFs...")
            
            def load_pdfs_thread():
                try:
                    self.pdf_paths = list(file_paths)
                    self.pdf_path = self.pdf_paths[0]  # Load first PDF for viewing
                    
                    # Update progress
                    total_files = len(self.pdf_paths)
                    for i, pdf_path in enumerate(self.pdf_paths):
                        progress = int((i / total_files) * 50)  # First 50% for file processing
                        self.after(0, lambda p=progress: self.update_loading_progress(p))
                        time.sleep(0.01)  # Small delay to show progress
                    
                    # Update UI elements on main thread
                    def update_ui():
                        self.file_label.configure(text=f"{len(self.pdf_paths)} PDFs selected")
                        self.file_count_label.configure(text=f"{len(self.pdf_paths)} files")
                        self.batch_btn.configure(state="normal")  # Enable batch process
                        self.apply_all_btn.configure(state="normal")  # Enable apply to all
                        self.status_label.configure(text=f"Selected {len(self.pdf_paths)} PDFs for batch processing")
                        self.export_btn.configure(state="disabled")
                        
                        # Clear previous extractions
                        self.extracted_sections = {}
                        self.file_previews = {}  # Clear previous previews
                        
                        # Initialize preview entries for all files
                        for pdf_path in self.pdf_paths:
                            file_name = Path(pdf_path).name
                            # Create empty placeholder - will be filled when selections are applied
                            self.file_previews[file_name] = pd.DataFrame()
                        
                        # Update progress to 100%
                        self.update_loading_progress(100)
                        time.sleep(0.2)
                        
                        # Update preview dropdown to show all files
                        self.update_excel_preview()
                        
                        # Load first PDF for viewing
                        self.load_pdf_viewer(self.pdf_paths[0])
                        
                        # Hide loading screen
                        self.hide_loading_screen()
                    
                    self.after(0, update_ui)
                    
                except Exception as e:
                    def show_error():
                        self.hide_loading_screen()
                        messagebox.showerror("Error", f"Failed to load PDFs:\n{str(e)}")
                    self.after(0, show_error)
            
            # Start thread
            thread = threading.Thread(target=load_pdfs_thread, daemon=True)
            thread.start()
    
    def select_folder(self):
        """Open dialog to select folder containing PDF files"""
        folder_path = filedialog.askdirectory(
            title="Select Folder Containing PDF Files"
        )
        
        if folder_path:
            self.show_loading_screen("Scanning folder...")
            
            def scan_folder_thread():
                try:
                    # Find all PDF files in the folder
                    self.after(0, lambda: self.update_loading_progress(25))
                    folder = Path(folder_path)
                    pdf_files = list(folder.glob("*.pdf"))
                    
                    self.after(0, lambda: self.update_loading_progress(75))
                    
                    if pdf_files:
                        self.pdf_paths = [str(f) for f in pdf_files]
                        self.pdf_path = self.pdf_paths[0]  # Load first PDF for viewing
                        
                        def update_ui():
                            self.file_label.configure(text=f"{len(self.pdf_paths)} PDFs found in folder")
                            self.file_count_label.configure(text=f"{len(self.pdf_paths)} files")
                            self.batch_btn.configure(state="normal")  # Enable batch process
                            self.apply_all_btn.configure(state="normal")  # Enable apply to all
                            self.status_label.configure(text=f"Found {len(self.pdf_paths)} PDFs in {folder.name}")
                            self.export_btn.configure(state="disabled")
                            
                            # Clear previous extractions
                            self.extracted_sections = {}
                            self.file_previews = {}  # Clear previous previews
                            
                            # Initialize preview entries for all files
                            for pdf_path in self.pdf_paths:
                                file_name = Path(pdf_path).name
                                # Create empty placeholder - will be filled when selections are applied
                                self.file_previews[file_name] = pd.DataFrame()
                            
                            # Update preview dropdown to show all files
                            self.update_excel_preview()
                            
                            # Load first PDF for viewing
                            self.load_pdf_viewer(self.pdf_paths[0])
                            
                            self.update_loading_progress(100)
                            time.sleep(0.2)
                            self.hide_loading_screen()
                        
                        self.after(0, update_ui)
                    else:
                        def show_no_files():
                            self.hide_loading_screen()
                            messagebox.showwarning("No PDFs Found", "No PDF files found in the selected folder.")
                            self.status_label.configure(text="No PDFs found in folder")
                        self.after(0, show_no_files)
                        
                except Exception as e:
                    def show_error():
                        self.hide_loading_screen()
                        messagebox.showerror("Error", f"Failed to scan folder:\n{str(e)}")
                    self.after(0, show_error)
            
            # Start thread
            thread = threading.Thread(target=scan_folder_thread, daemon=True)
            thread.start()
    
    def is_header_footer(self, text):
        """Check if text is part of header or footer"""
        if not text or pd.isna(text):
            return True
        
        text_str = str(text).strip()
        text_lower = text_str.lower()
        
        # Exact footer patterns
        footer_keywords = [
            'commercial confidential',
            'experian information services (malaysia) sdn. bhd.',
            'is certified to iso/iec',
            'cert. no: ism',
            'notice: the information provided by experian',
            'we do not guarantee the accuracy',
            'while we have used our best endeavours',
            'the information furnished is strictly confidential',
            'experian shall not be liable',
            'customer service division at:',
            'suite 16.02, level 16',
            'centrepoint south mid valley city',
            'lingkaran syed putra',
            'kuala lumpur',
            '+60326151111',
            'page',
            'of 7'
        ]
        
        # Exact header patterns
        header_keywords = [
            'strictly confidential',
            'order id',
            'credittrack by experian',
            'order date',
            'effective date',
            'user name',
            'check/track by experian'
        ]
        
        # Check footer patterns
        for keyword in footer_keywords:
            if keyword in text_lower:
                return True
        
        # Check header patterns
        for keyword in header_keywords:
            if keyword in text_lower:
                return True
        
        # Check if text is mostly the disclaimer
        if len(text_str) > 200 and 'experian' in text_lower:
            return True
        
        # Check if it's just a page number pattern
        if re.match(r'^page\s+\d+\s+of\s+\d+$', text_lower):
            return True
        
        # Check for Order ID pattern
        if re.match(r'.*order\s+id\s*:\s*\d+.*', text_lower):
            return True
            
        return False
    
    def clean_table_data(self, df):
        """Remove header/footer rows from dataframe"""
        if df.empty:
            return df
        
        # Remove rows where ANY column contains header/footer text
        rows_to_keep = []
        for idx, row in df.iterrows():
            # Check if this row contains header/footer content
            is_bad_row = False
            for cell in row:
                if self.is_header_footer(cell):
                    is_bad_row = True
                    break
            
            if not is_bad_row:
                rows_to_keep.append(idx)
        
        df_cleaned = df.loc[rows_to_keep].reset_index(drop=True)
        
        # Remove columns that are all empty or None
        df_cleaned = df_cleaned.dropna(axis=1, how='all')
        
        return df_cleaned
    
    def clean_extracted_data(self):
        """Deep clean extracted data to remove all header/footer contamination"""
        if not self.extracted_sections:
            messagebox.showwarning("Warning", "No data to clean! Please extract data first.")
            return
        
        try:
            self.status_label.configure(text="Cleaning extracted data...")
            self.update()
            
            df = self.extracted_sections["Extracted Data"]
            original_rows = len(df)
            
            # Apply thorough cleaning
            df_cleaned = self.clean_table_data(df)
            
            # Update the section with cleaned data
            self.extracted_sections["Extracted Data"] = df_cleaned
            
            rows_removed = original_rows - len(df_cleaned)
            
            self.status_label.configure(text=f"✓ Cleaned! Removed {rows_removed} rows")
            
            # Update Excel preview
            self.update_excel_preview()
            
            messagebox.showinfo("Success", f"Data cleaned successfully!\n\nRows removed: {rows_removed}\nTotal rows: {len(df_cleaned)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to clean data:\n{str(e)}")
            self.status_label.configure(text="✗ Cleaning failed")
            import traceback
            traceback.print_exc()
    
    def load_pdf_viewer(self, pdf_path):
        """Load PDF into the PDF viewer tab"""
        try:
            if self.pdf_document:
                self.pdf_document.close()
            
            self.pdf_document = fitz.open(pdf_path)
            self.current_page = 0
            
            # Enable navigation buttons
            self.select_table_btn.configure(state="normal")
            self.save_selection_btn.configure(state="normal")
            self.clear_selections_btn.configure(state="normal")
            self.auto_extract_btn.configure(state="normal")
            
            # Enable apply button if multiple files selected
            if self.pdf_paths and len(self.pdf_paths) > 1:
                self.apply_all_btn.configure(state="normal")
            # Enable sticky overlay nav buttons
            try:
                self.sticky_prev.configure(state="normal")
                self.sticky_next.configure(state="normal")
            except Exception:
                pass
            # Enable clear preview if previews exist with data
            try:
                has_data = any((not df.empty) for df in self.file_previews.values()) if self.file_previews else False
                if has_data:
                    self.clear_preview_btn.configure(state="normal")
                else:
                    self.clear_preview_btn.configure(state="disabled")
            except Exception:
                pass
            
            # Display first page
            self.display_pdf_page()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load PDF viewer:\n{str(e)}")
    
    def display_pdf_page(self):
        """Display the current PDF page"""
        if not self.pdf_document:
            return
        
        try:
            page = self.pdf_document[self.current_page]
            
            # Update page label and sticky center
            page_text = f"Page {self.current_page + 1} of {len(self.pdf_document)}"
            self.page_label.configure(text=page_text)
            self.sticky_center.configure(text=page_text)
            
            # Render page to image
            zoom = 1.5  # Zoom factor for better quality
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # Convert to PIL Image
            img_data = pix.tobytes("ppm")
            img = Image.open(io.BytesIO(img_data))
            
            # Convert to PhotoImage
            photo = ImageTk.PhotoImage(img)
            
            # Clear canvas
            self.pdf_canvas.delete("all")
            
            # Get canvas size
            canvas_width = self.pdf_canvas.winfo_width()
            canvas_height = self.pdf_canvas.winfo_height()
            
            # Calculate center position
            img_width, img_height = img.size
            x_center = max(0, (canvas_width - img_width) // 2) if canvas_width > img_width else 0
            y_center = 0  # Keep top aligned for scrolling
            
            # Store offsets for coordinate conversion
            self.pdf_x_offset = x_center
            self.pdf_y_offset = y_center
            
            # Display image centered horizontally
            self.pdf_canvas.create_image(x_center, y_center, anchor="nw", image=photo)
            self.pdf_canvas.image = photo  # Keep a reference
            
            # Update scroll region
            self.pdf_canvas.configure(scrollregion=self.pdf_canvas.bbox("all"))
            
            # Redraw saved selections
            self.redraw_selections()
            
        except Exception as e:
            print(f"Error displaying PDF page: {e}")
    
    def previous_page(self):
        """Go to previous PDF page"""
        if self.current_page > 0:
            self.current_page -= 1
            self.display_pdf_page()
    
    def next_page(self):
        """Go to next PDF page"""
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display_pdf_page()
    
    def on_mouse_wheel(self, event):
        """Handle mouse wheel scrolling"""
        # Scroll the canvas vertically
        self.pdf_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    def toggle_selection_mode(self):
        """Toggle table selection mode"""
        self.selecting = not self.selecting
        if self.selecting:
            self.select_table_btn.configure(text="✓ Selection Mode ON", fg_color="green")
            self.status_label.configure(text="📐 Draw a rectangle around the table you want to extract")
        else:
            self.select_table_btn.configure(text="📐 Select Table Area", fg_color="purple")
            self.status_label.configure(text="Selection mode disabled")
    
    def on_mouse_down(self, event):
        """Start drawing selection rectangle"""
        if not self.selecting:
            return
        
        # Convert event coordinates to canvas coordinates (account for scroll)
        canvas_x = self.pdf_canvas.canvasx(event.x)
        canvas_y = self.pdf_canvas.canvasy(event.y)
        self.selection_start = (canvas_x, canvas_y)
        # Delete previous rectangle if exists
        if self.selection_rect:
            self.pdf_canvas.delete(self.selection_rect)
            self.selection_rect = None
    
    def on_mouse_drag(self, event):
        """Draw selection rectangle as user drags"""
        if not self.selecting or not self.selection_start:
            return
        
        # Delete previous rectangle
        if self.selection_rect:
            self.pdf_canvas.delete(self.selection_rect)
        
        # Draw new rectangle
        x1, y1 = self.selection_start
        # Convert event coordinates to canvas coordinates (account for scroll)
        x2 = self.pdf_canvas.canvasx(event.x)
        y2 = self.pdf_canvas.canvasy(event.y)
        
        self.selection_rect = self.pdf_canvas.create_rectangle(
            x1, y1, x2, y2,
            outline="red",
            width=3,
            dash=(5, 5)
        )
    
    def on_mouse_up(self, event):
        """Finish selection and store temporarily until saved"""
        if not self.selecting or not self.selection_start:
            return
        
        x1, y1 = self.selection_start
        # Convert event coordinates to canvas coordinates (account for scroll)
        x2 = self.pdf_canvas.canvasx(event.x)
        y2 = self.pdf_canvas.canvasy(event.y)
        
        # Ensure x1 < x2 and y1 < y2
        if x1 > x2:
            x1, x2 = x2, x1
        if y1 > y2:
            y1, y2 = y2, y1
        
        # Convert canvas coordinates to PDF coordinates
        # First subtract the PDF offset (centering)
        pdf_x1 = x1 - self.pdf_x_offset
        pdf_y1 = y1 - self.pdf_y_offset
        pdf_x2 = x2 - self.pdf_x_offset
        pdf_y2 = y2 - self.pdf_y_offset
        
        # Then account for zoom factor
        zoom = 1.5
        
        # Get bounding box in PDF coordinates
        bbox = (pdf_x1/zoom, pdf_y1/zoom, pdf_x2/zoom, pdf_y2/zoom)
        
        # Draw selection rectangle (orange for temporary)
        perm_rect = self.pdf_canvas.create_rectangle(
            x1, y1, x2, y2,
            outline="orange",
            width=2
        )
        
        # Store in temporary selections
        if self.current_page not in self.temp_selections:
            self.temp_selections[self.current_page] = []
        self.temp_selections[self.current_page].append((bbox, perm_rect))
        
        # Add to history for undo
        self.selection_history.append((self.current_page, bbox, perm_rect))
        
        # Clear temporary selection
        if self.selection_rect:
            self.pdf_canvas.delete(self.selection_rect)
            self.selection_rect = None
        
        self.selection_start = None
        
        # Count temp selections on current page
        temp_count = len(self.temp_selections.get(self.current_page, []))
        self.status_label.configure(text=f"✓ Selection marked! Page {self.current_page + 1}: {temp_count} selection(s) (Click 💾 Save to finalize)")
    
    def save_current_page_selections(self):
        """Save all temporary selections for current page and extract data with formatting"""
        if self.current_page not in self.temp_selections or not self.temp_selections[self.current_page]:
            messagebox.showinfo("Info", "No selections to save on this page!")
            return
        
        self.show_loading_screen("Saving selections...")
        
        def save_selections_thread():
            try:
                self.after(0, lambda: self.update_loading_progress(25))
                
                # Get current viewing file (could be different from batch files)
                current_file = self.pdf_path
                
                # Extract i-SCORE and risk grade from current file
                file_name = Path(current_file).name
                if file_name not in self.file_risk_grades:
                    iscore, risk_grade = self.extract_iscore_and_risk_grade(current_file)
                    if risk_grade:
                        self.file_risk_grades[file_name] = risk_grade
                
                # Extract data from all temp selections on this page
                with pdfplumber.open(current_file) as pdf:
                    page = pdf.pages[self.current_page]
                    
                    selection_count = len(self.temp_selections[self.current_page])
                    
                    for idx, (bbox, rect_id) in enumerate(self.temp_selections[self.current_page]):
                        # Update progress for each selection
                        progress = 25 + int((idx / selection_count) * 50)
                        self.after(0, lambda p=progress: self.update_loading_progress(p))
                        
                        # Extract table data
                        cropped = page.crop(bbox)
                        table = cropped.extract_table()
                        
                        if table:
                            new_df = pd.DataFrame(table)
                            
                            # Append to file preview
                            if file_name in self.file_previews:
                                existing_df = self.file_previews[file_name]
                                # Add separator
                                num_cols = max(len(existing_df.columns), len(new_df.columns))
                                separator = pd.DataFrame([[None] * num_cols] * 2)
                                self.file_previews[file_name] = pd.concat([existing_df, separator, new_df], ignore_index=True)
                            else:
                                self.file_previews[file_name] = new_df
                        
                        # Change rectangle color to blue (saved) on main thread
                        self.after(0, lambda rid=rect_id: self.pdf_canvas.itemconfig(rid, outline="blue"))
                        
                        # Move to saved selections
                        if self.current_page not in self.saved_selections:
                            self.saved_selections[self.current_page] = []
                        self.saved_selections[self.current_page].append(bbox)
                
                # Clear temp selections for this page
                saved_count = len(self.temp_selections[self.current_page])
                self.temp_selections[self.current_page] = []
                
                def finish_save():
                    self.update_loading_progress(100)
                    time.sleep(0.2)
                    
                    # Update preview
                    self.update_excel_preview()
                    self.export_btn.configure(state="normal")
                    
                    total_saved = sum(len(sels) for sels in self.saved_selections.values())
                    self.status_label.configure(text=f"✓ Saved {saved_count} selection(s) from page {self.current_page + 1}! Total saved: {total_saved}")
                    
                    self.hide_loading_screen()
                
                self.after(0, finish_save)
                
            except Exception as e:
                def show_error():
                    self.hide_loading_screen()
                    messagebox.showerror("Error", f"Failed to save selections:\n{str(e)}")
                    print(f"Save error: {e}")
                self.after(0, show_error)
        
        # Start thread
        thread = threading.Thread(target=save_selections_thread, daemon=True)
        thread.start()
    
    def undo_last_selection(self):
        """Undo the most recent selection (remove from temp selections)"""
        if not self.selection_history:
            messagebox.showinfo("Info", "No selections to undo!")
            return
        
        # Get last selection from history
        page_num, bbox, rect_id = self.selection_history.pop()
        
        # Remove from temp selections
        if page_num in self.temp_selections:
            self.temp_selections[page_num] = [(b, r) for b, r in self.temp_selections[page_num] if r != rect_id]
            if not self.temp_selections[page_num]:
                del self.temp_selections[page_num]
        
        # Delete rectangle from canvas
        self.pdf_canvas.delete(rect_id)
        
        # Update status
        remaining = sum(len(sels) for sels in self.temp_selections.values())
        self.status_label.configure(text=f"↶ Undone! {remaining} unsaved selection(s) remaining")
    
    def auto_extract_all_tables(self):
        """Automatically extract all tables from all pages with header/footer removal"""
        # Auto-extract should run across all selected PDFs (batch) or the single loaded PDF
        files = self.pdf_paths if self.pdf_paths else ([self.pdf_path] if self.pdf_path else [])
        if not files:
            messagebox.showinfo("Info", "No PDF(s) loaded!")
            return

        self.show_loading_screen("Auto-extracting tables...")
        
        def auto_extract_thread():
            try:
                header_crop = 90  # Top 80px to remove
                footer_crop = 90  # Bottom 80px to remove
                processed = 0
                total_files = len(files)

                for file_idx, current_file in enumerate(files):
                    try:
                        file_name = Path(current_file).name
                        
                        # Extract i-SCORE and risk grade
                        if file_name not in self.file_risk_grades:
                            iscore, risk_grade = self.extract_iscore_and_risk_grade(current_file)
                            if risk_grade:
                                self.file_risk_grades[file_name] = risk_grade
                        
                        # Update progress
                        progress = int((file_idx / total_files) * 90)
                        self.after(0, lambda p=progress, fn=file_name: (
                            self.update_loading_progress(p),
                            self.status_label.configure(text=f"Auto-extracting tables from {fn}...")
                        ))

                        with pdfplumber.open(current_file) as pdf:
                            all_tables = []
                            
                            for page in pdf.pages:
                                page_height = page.height
                                cropped_page = page.crop((0, header_crop, page.width, page_height - footer_crop))
                                tables = cropped_page.extract_tables()
                                
                                if tables:
                                    # Get table objects with cell bounding boxes for color detection
                                    table_settings = {
                                        "vertical_strategy": "lines",
                                        "horizontal_strategy": "lines",
                                        "snap_tolerance": 3,
                                        "join_tolerance": 3,
                                    }
                                    tables_with_cells = cropped_page.find_tables(table_settings)
                                    
                                    for table_idx, table in enumerate(tables):
                                        if table and len(table) > 0:
                                            all_tables.append(table)

                            if all_tables:
                                combined_rows = []
                                for table_idx, table in enumerate(all_tables):
                                    combined_rows.extend(table)
                                    if table_idx < len(all_tables) - 1:
                                        num_cols = len(table[0]) if table else 1
                                        combined_rows.extend([[None] * num_cols] * 2)

                                df = pd.DataFrame(combined_rows)
                                self.file_previews[file_name] = df
                                processed += 1
                    except Exception as e:
                        # Continue to next file if one fails
                        print(f"Auto-extract failed for {current_file}: {e}")
                        continue

                # Final UI updates
                def finish_extraction():
                    self.update_loading_progress(100)
                    time.sleep(0.3)
                    
                    if processed > 0:
                        self.update_excel_preview()
                        self.export_btn.configure(state="normal")
                        self.clear_preview_btn.configure(state="normal")
                        self.status_label.configure(text=f"✓ Auto-extracted tables from {processed}/{len(files)} files")
                        self.hide_loading_screen()
                        messagebox.showinfo("Success", f"Auto-extracted tables from {processed}/{len(files)} files.\n\nCheck Excel Preview tab.")
                    else:
                        self.hide_loading_screen()
                        messagebox.showwarning("No Tables", "No tables found in the selected PDF(s).")
                        self.status_label.configure(text="⚠ No tables found")

                self.after(0, finish_extraction)

            except Exception as e:
                def show_error():
                    self.hide_loading_screen()
                    messagebox.showerror("Error", f"Auto-extraction failed:\n{str(e)}")
                    self.status_label.configure(text="✗ Auto-extraction failed")
                    import traceback
                    traceback.print_exc()
                self.after(0, show_error)

        # Start thread
        thread = threading.Thread(target=auto_extract_thread, daemon=True)
        thread.start()
    
    def apply_selections_to_all_files(self):
        """Apply saved selections to all imported PDF files with header/footer removal"""
        if not self.pdf_paths:
            messagebox.showinfo("Info", "No multiple files selected!")
            return
        
        if not self.saved_selections:
            messagebox.showerror("Error", "No selections saved! Please mark and save selections first.")
            return
        
        self.show_loading_screen("Applying selections to all files...")
        
        def apply_selections_thread():
            try:
                processed_count = 0
                header_crop = 80  # Top 80px to remove
                footer_crop = 80  # Bottom 80px to remove
                total_files = len(self.pdf_paths)
                
                for file_idx, pdf_path in enumerate(self.pdf_paths):
                    file_name = Path(pdf_path).name
                    
                    # Extract i-SCORE and risk grade
                    if file_name not in self.file_risk_grades:
                        iscore, risk_grade = self.extract_iscore_and_risk_grade(pdf_path)
                        if risk_grade:
                            self.file_risk_grades[file_name] = risk_grade
                    
                    # Update progress
                    progress = int((file_idx / total_files) * 90)
                    self.after(0, lambda p=progress, fn=file_name: (
                        self.update_loading_progress(p),
                        self.status_label.configure(text=f"Applying selections to {fn}...")
                    ))
                    
                    try:
                        with pdfplumber.open(pdf_path) as pdf:
                            file_tables = []
                            
                            # Apply selections from each saved page
                            for page_num in sorted(self.saved_selections.keys()):
                                if page_num < len(pdf.pages):
                                    page = pdf.pages[page_num]
                                    page_height = page.height
                                    
                                    # Crop page to remove header/footer
                                    cropped_page = page.crop((0, header_crop, page.width, page_height - footer_crop))
                                    
                                    for bbox in self.saved_selections[page_num]:
                                        try:
                                            # Adjust bbox coordinates for the cropped page
                                            adjusted_bbox = (
                                                bbox[0],
                                                max(0, bbox[1] - header_crop),  # Adjust Y coordinate
                                                bbox[2],
                                                bbox[3] - header_crop
                                            )
                                            
                                            # Only extract if bbox is within cropped area
                                            if adjusted_bbox[1] >= 0 and adjusted_bbox[3] <= (page_height - header_crop - footer_crop):
                                                table = cropped_page.crop(adjusted_bbox).extract_table()
                                                if table:
                                                    file_tables.append(pd.DataFrame(table))
                                        except:
                                            continue
                            
                            # Combine all tables for this file
                            if file_tables:
                                combined_df = file_tables[0]
                                for df in file_tables[1:]:
                                    # Add separator
                                    num_cols = max(len(combined_df.columns), len(df.columns))
                                    separator = pd.DataFrame([[None] * num_cols] * 2)
                                    combined_df = pd.concat([combined_df, separator, df], ignore_index=True)
                                
                                self.file_previews[file_name] = combined_df
                                processed_count += 1
                    
                    except Exception as e:
                        print(f"Error processing {file_name}: {e}")
                        continue
                
                def finish_apply():
                    self.update_loading_progress(100)
                    time.sleep(0.3)
                    
                    # Update preview
                    self.update_excel_preview()
                    self.export_btn.configure(state="normal")
                    
                    self.status_label.configure(text=f"✓ Applied to {processed_count} files with header/footer removed!")
                    self.hide_loading_screen()
                    messagebox.showinfo("Success", f"Successfully applied selections to {processed_count} files!\n\nHeader/footer (80px) removed automatically.\n\nCheck the Excel Preview tab.")
                
                self.after(0, finish_apply)
                
            except Exception as e:
                def show_error():
                    self.hide_loading_screen()
                    messagebox.showerror("Error", f"Failed to apply selections:\n{str(e)}")
                    print(f"Apply error: {e}")
                self.after(0, show_error)
        
        # Start thread
        thread = threading.Thread(target=apply_selections_thread, daemon=True)
        thread.start()
    
    def redraw_selections(self):
        """Redraw saved and temporary selections for the current page only"""
        # Note: We don't clear rectangles here as they persist across page changes
        # Only redraw if switching pages
        
        zoom = 1.5
        
        # Redraw saved selections (blue)
        if self.current_page in self.saved_selections:
            for bbox in self.saved_selections[self.current_page]:
                x1, y1, x2, y2 = bbox
                # Convert PDF coords to canvas coords and add offset
                canvas_x1 = x1 * zoom + self.pdf_x_offset
                canvas_y1 = y1 * zoom + self.pdf_y_offset
                canvas_x2 = x2 * zoom + self.pdf_x_offset
                canvas_y2 = y2 * zoom + self.pdf_y_offset
                
                self.pdf_canvas.create_rectangle(
                    canvas_x1, canvas_y1, canvas_x2, canvas_y2,
                    outline="blue",
                    width=2
                )
        
        # Redraw temp selections (orange)
        if self.current_page in self.temp_selections:
            for bbox, old_rect_id in self.temp_selections[self.current_page]:
                x1, y1, x2, y2 = bbox
                # Convert PDF coords to canvas coords and add offset
                canvas_x1 = x1 * zoom + self.pdf_x_offset
                canvas_y1 = y1 * zoom + self.pdf_y_offset
                canvas_x2 = x2 * zoom + self.pdf_x_offset
                canvas_y2 = y2 * zoom + self.pdf_y_offset
                
                self.pdf_canvas.create_rectangle(
                    canvas_x1, canvas_y1, canvas_x2, canvas_y2,
                    outline="orange",
                    width=2
                )
    
    def extract_selected_table(self, bbox):
        """Extract table from the selected bounding box"""
        if not self.pdf_document:
            return
        
        try:
            self.status_label.configure(text="Extracting table from selected area...")
            self.update()
            
            with pdfplumber.open(self.pdf_path) as pdf:
                page = pdf.pages[self.current_page]
                
                # Extract table from the selected area
                table = page.crop(bbox).extract_table()
                
                if table:
                    # Convert to DataFrame
                    new_df = pd.DataFrame(table)
                    
                    # Append to single sheet named "Extracted Data"
                    if "Extracted Data" in self.extracted_sections:
                        # Append to existing data with blank rows as separator
                        existing_df = self.extracted_sections["Extracted Data"]
                        
                        # Create separator (2 blank rows)
                        num_cols = max(len(existing_df.columns), len(new_df.columns))
                        separator = pd.DataFrame([[None] * num_cols] * 2)
                        
                        # Concatenate: existing data + separator + new data
                        self.extracted_sections["Extracted Data"] = pd.concat([existing_df, separator, new_df], ignore_index=True)
                    else:
                        # Create new sheet
                        self.extracted_sections["Extracted Data"] = new_df
                    
                    # Enable export buttons
                    self.export_btn.configure(state="normal")
                    
                    # Update displays
                    self.update_excel_preview()
                    
                    # Show success message
                    total_rows = len(self.extracted_sections["Extracted Data"])
                    self.status_label.configure(text=f"✓ Table extracted! Total: {total_rows} rows")
                    
                    # Turn off selection mode
                    self.selecting = False
                    self.select_table_btn.configure(text="📐 Select Table Area", fg_color="purple")
                    
                    # Clear selection rectangle
                    if self.selection_rect:
                        self.pdf_canvas.delete(self.selection_rect)
                        self.selection_rect = None
                else:
                    messagebox.showwarning("No Table", "No table found in selected area. Try selecting a different area.")
                    self.status_label.configure(text="⚠ No table found in selected area")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to extract table:\n{str(e)}")
            self.status_label.configure(text="✗ Extraction failed")
            import traceback
            traceback.print_exc()
    
    def update_excel_preview(self):
        """Update Excel preview tab with extracted data"""
        if not self.file_previews:
            self.sheet_selector.configure(values=["No data available"], state="disabled")
            # Clear treeview
            for item in self.excel_display.get_children():
                self.excel_display.delete(item)
            try:
                self.clear_preview_btn.configure(state="disabled")
            except Exception:
                pass
            return
        
        # Show list of files
        file_names = list(self.file_previews.keys())
        self.sheet_selector.configure(values=file_names, state="readonly")
        self.sheet_selector.set(file_names[0])
        
        # Display first file
        self.display_file_preview(file_names[0])
        # Enable Clear Preview if any file has data
        try:
            has_data = any((not df.empty) for df in self.file_previews.values())
            if has_data:
                self.clear_preview_btn.configure(state="normal")
            else:
                self.clear_preview_btn.configure(state="disabled")
        except Exception:
            pass
        
        # Update database file list
        try:
            self.update_database_file_list()
        except Exception:
            pass
    
    def on_file_selected(self, file_name):
        """Handle file selection change"""
        self.display_file_preview(file_name)

    def clear_all_previews(self):
        """Clear all Excel preview data (reset preview dropdown and display)"""
        self.show_loading_screen("Clearing previews...")
        
        def clear_previews_thread():
            try:
                self.after(0, lambda: self.update_loading_progress(25))
                
                # Clear file previews
                self.file_previews = {}
                
                self.after(0, lambda: self.update_loading_progress(50))
                
                def update_ui():
                    # Reset selector and display
                    try:
                        self.sheet_selector.configure(values=["No data available"], state="disabled")
                    except Exception:
                        pass
                    
                    self.update_loading_progress(75)
                    
                    # Clear treeview
                    for item in self.excel_display.get_children():
                        self.excel_display.delete(item)
                    
                    # Disable export and clear-preview button
                    try:
                        self.export_btn.configure(state="disabled")
                        self.clear_preview_btn.configure(state="disabled")
                    except Exception:
                        pass
                    
                    self.update_loading_progress(100)
                    time.sleep(0.2)
                    
                    self.status_label.configure(text="Excel preview cleared")
                    self.hide_loading_screen()
                
                self.after(0, update_ui)
                
            except Exception as e:
                def show_error():
                    self.hide_loading_screen()
                    messagebox.showerror("Error", f"Failed to clear previews:\n{str(e)}")
                self.after(0, show_error)
        
        # Start thread
        thread = threading.Thread(target=clear_previews_thread, daemon=True)
        thread.start()
    
    def display_file_preview(self, file_name):
        """Display selected file preview data in Excel-like grid"""
        if file_name not in self.file_previews:
            return
        
        df = self.file_previews[file_name]
        
        # Clear existing treeview
        for item in self.excel_display.get_children():
            self.excel_display.delete(item)
        
        # Clear existing columns
        self.excel_display["columns"] = []
        
        # Check if dataframe is empty
        if df.empty:
            # Show message in treeview
            self.excel_display["columns"] = ["Message"]
            self.excel_display.heading("#0", text="")
            self.excel_display.column("#0", width=0, stretch=False)
            self.excel_display.heading("Message", text=f"File: {file_name}")
            self.excel_display.column("Message", width=800, anchor="w")
            
            self.excel_display.insert("", "end", values=("No data extracted yet.",))
            self.excel_display.insert("", "end", values=("",))
            self.excel_display.insert("", "end", values=("To extract data:",))
            self.excel_display.insert("", "end", values=("1. Mark selection areas on PDF",))
            self.excel_display.insert("", "end", values=("2. Click 💾 Save Selections",))
            self.excel_display.insert("", "end", values=("3. Click 📋 Apply to All Files to process all PDFs",))
        else:
            # Set up columns - use Excel-style letters (A, B, C, etc.)
            num_cols = len(df.columns)
            col_letters = [self.excel_column_letter(i) for i in range(num_cols)]
            
            self.excel_display["columns"] = col_letters
            
            # Hide the first column (tree column)
            self.excel_display.heading("#0", text="Row")
            self.excel_display.column("#0", width=50, anchor="center")
            
            # Configure column headers with improved layout
            standard_width = 150  # Equal width for all columns, same as database preview
            for i, col_letter in enumerate(col_letters):
                self.excel_display.heading(col_letter, text=col_letter, anchor='w')
                self.excel_display.column(col_letter, width=standard_width, minwidth=100, anchor='w')
            
            # Add data rows with improved formatting
            for idx, row in df.iterrows():
                # Format values - replace empty/NaN with "-" for consistency
                row_values = []
                for val in row:
                    if pd.isna(val) or val == "" or str(val).strip() == "":
                        row_values.append("-")
                    else:
                        # Clean and format cell value
                        formatted_value = str(val).strip()
                        # Limit very long values to prevent display issues
                        if len(formatted_value) > 50:
                            formatted_value = formatted_value[:47] + "..."
                        row_values.append(formatted_value)
                
                self.excel_display.insert("", "end", text=str(idx + 1), values=row_values)
    
    def excel_column_letter(self, n):
        """Convert column number to Excel-style letter (0->A, 1->B, 25->Z, 26->AA, etc.)"""
        result = ""
        while n >= 0:
            result = chr(n % 26 + 65) + result
            n = n // 26 - 1
            if n < 0:
                break
        return result
    
    def extract_data(self):
        """Extract ALL tables from PDF directly - simple and accurate"""
        if not self.pdf_path:
            messagebox.showerror("Error", "Please select a PDF file first!")
            return
        
        try:
            self.status_label.configure(text="Extracting data from PDF...")
            self.update()
            
            self.extracted_sections = {}
            table_counter = 1
            
            with pdfplumber.open(self.pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    # Extract ALL tables from this page
                    tables = page.extract_tables()
                    
                    if tables:
                        for table in tables:
                            if not table or len(table) < 1:
                                continue
                            
                            # Convert to DataFrame - keep everything as-is
                            df = pd.DataFrame(table)
                            
                            if not df.empty:
                                # Simple naming: Table_1, Table_2, etc.
                                section_name = f"Table_{table_counter}"
                                
                                # Try to identify section from first row
                                try:
                                    first_row = ' '.join([str(cell) for cell in df.iloc[0] if pd.notna(cell)])
                                    
                                    if "SECTION 1" in first_row.upper():
                                        section_name = "Section 1"
                                    elif "SECTION 2" in first_row.upper():
                                        section_name = "Section 2"
                                    elif "SECTION 3" in first_row.upper():
                                        section_name = "Section 3"
                                    elif "SECTION 4" in first_row.upper():
                                        section_name = "Section 4"
                                    elif "SECTION 5" in first_row.upper():
                                        section_name = "Section 5"
                                except:
                                    pass
                                
                                self.extracted_sections[section_name] = df
                                table_counter += 1
            
            # Display extracted sections
            self.text_display.delete("1.0", "end")
            self.text_display.insert("1.0", f"✓ Extraction Complete!\n")
            self.text_display.insert("end", f"{'='*100}\n\n")
            self.text_display.insert("end", f"Total Sections Found: {len(self.extracted_sections)}\n\n")
            
            for idx, (section_name, df) in enumerate(self.extracted_sections.items(), 1):
                self.text_display.insert("end", f"\n{'='*100}\n")
                self.text_display.insert("end", f"SECTION {idx}: {section_name}\n")
                self.text_display.insert("end", f"{'='*100}\n")
                self.text_display.insert("end", f"Rows: {len(df)} | Columns: {len(df.columns)}\n")
                self.text_display.insert("end", f"Columns: {list(df.columns)}\n\n")
                
                # Show first 10 rows
                preview = df.head(10).to_string()
                self.text_display.insert("end", preview + "\n")
                
                if len(df) > 10:
                    self.text_display.insert("end", f"\n... and {len(df) - 10} more rows\n")
            
            # Enable export button
            if self.extracted_sections:
                self.export_btn.configure(state="normal")
                self.clean_btn.configure(state="normal")
                self.status_label.configure(text=f"✓ Extracted {len(self.extracted_sections)} sections successfully!")
                
                # Update Excel preview
                self.update_excel_preview()
                
                messagebox.showinfo("Success", f"Successfully extracted {len(self.extracted_sections)} sections from the PDF!")
            else:
                self.status_label.configure(text="⚠ No sections found in PDF")
                messagebox.showwarning("Warning", "No recognizable sections found in the PDF!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to extract data:\n{str(e)}")
            self.status_label.configure(text="✗ Extraction failed")
            self.text_display.delete("1.0", "end")
            self.text_display.insert("1.0", f"Error: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def batch_process_pdfs(self):
        """Process multiple PDFs using saved selection areas with header/footer removal"""
        if not self.pdf_paths:
            messagebox.showerror("Error", "No PDF files selected for batch processing!")
            return
        
        if not self.saved_selections:
            messagebox.showerror("Error", "No selection areas defined! Please mark selection areas first.")
            return
        
        # Ask user to select output folder
        output_folder = filedialog.askdirectory(
            title="Select Output Folder for Excel Files"
        )
        
        if not output_folder:
            return
        
        output_path = Path(output_folder)
        success_count = 0
        failed_files = []
        header_crop = 80  # Top 80px to remove
        footer_crop = 80  # Bottom 80px to remove
        
        try:
            total_files = len(self.pdf_paths)
            
            for idx, pdf_path in enumerate(self.pdf_paths, 1):
                try:
                    pdf_name = Path(pdf_path).stem
                    self.status_label.configure(text=f"Processing {idx}/{total_files}: {pdf_name}...")
                    self.update()
                    
                    # Extract data from this PDF using saved selections
                    with pdfplumber.open(pdf_path) as pdf:
                        all_extracted_tables = []
                        
                        # Sort pages to maintain order
                        sorted_pages = sorted(self.saved_selections.keys())
                        
                        # Apply selections from each page in order
                        for page_num in sorted_pages:
                            if page_num < len(pdf.pages):
                                page = pdf.pages[page_num]
                                page_height = page.height
                                
                                # Crop page to remove header/footer
                                cropped_page = page.crop((0, header_crop, page.width, page_height - footer_crop))
                                
                                for bbox in self.saved_selections[page_num]:
                                    try:
                                        # Adjust bbox coordinates for the cropped page
                                        adjusted_bbox = (
                                            bbox[0],
                                            max(0, bbox[1] - header_crop),
                                            bbox[2],
                                            bbox[3] - header_crop
                                        )
                                        
                                        # Only extract if bbox is within cropped area
                                        if adjusted_bbox[1] >= 0 and adjusted_bbox[3] <= (page_height - header_crop - footer_crop):
                                            table = cropped_page.crop(adjusted_bbox).extract_table()
                                            if table:
                                                all_extracted_tables.append(table)
                                    except:
                                        continue
                        
                        if all_extracted_tables:
                            # Combine all tables with separators
                            combined_rows = []
                            for table_idx, table in enumerate(all_extracted_tables):
                                if table:
                                    combined_rows.extend(table)
                                    # Add separator between tables
                                    if table_idx < len(all_extracted_tables) - 1:
                                        num_cols = len(table[0]) if table else 1
                                        combined_rows.extend([[None] * num_cols] * 2)
                            
                            df = pd.DataFrame(combined_rows)
                            
                            # Export to Excel
                            excel_path = output_path / f"{pdf_name}.xlsx"
                            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                                df.to_excel(writer, sheet_name="Extracted Data", index=False, header=False)
                            
                            success_count += 1
                        else:
                            failed_files.append(f"{pdf_name} (no tables found)")
                
                except Exception as e:
                    failed_files.append(f"{Path(pdf_path).name} (error: {str(e)})")
                    continue
            
            # Show summary
            total_selections = sum(len(sels) for sels in self.saved_selections.values())
            summary_msg = f"Batch Processing Complete!\n\n"
            summary_msg += f"Successfully processed: {success_count}/{total_files}\n"
            summary_msg += f"Used {total_selections} selection areas across {len(self.saved_selections)} pages\n"
            summary_msg += f"Header/footer (80px) removed automatically\n"
            summary_msg += f"Output folder: {output_folder}\n"
            
            if failed_files:
                summary_msg += f"\nFailed files ({len(failed_files)}): \n"
                summary_msg += "\n".join(failed_files[:5])
                if len(failed_files) > 5:
                    summary_msg += f"\n... and {len(failed_files) - 5} more"
            
            self.status_label.configure(text=f"✓ Batch complete: {success_count}/{total_files} successful")
            messagebox.showinfo("Batch Processing Complete", summary_msg)
            
        except Exception as e:
            messagebox.showerror("Batch Processing Error", f"An error occurred:\n{str(e)}")
            self.status_label.configure(text="✗ Batch processing failed")
            import traceback
            traceback.print_exc()
    
    def export_to_excel(self):
        """Export all extracted tables to Excel with thick borders, bold detection, and merged headers"""
        if not self.file_previews:
            messagebox.showerror("Error", "No data to export! Please mark areas first.")
            return
        
        try:
            # Open save dialog
            save_folder = filedialog.askdirectory(
                title="Select Output Folder for Excel Files"
            )
            
            if save_folder:
                self.status_label.configure(text="Exporting to Excel...")
                self.update()
                
                from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
                from openpyxl.utils import get_column_letter
                
                output_path = Path(save_folder)
                
                # Export each file preview
                for file_name, df in self.file_previews.items():
                    excel_name = Path(file_name).stem + ".xlsx"
                    excel_path = output_path / excel_name
                    
                    # Clean dataframe before export - PRESERVE TABLE STRUCTURE
                    df_cleaned = df.copy()
                    
                    # Step 1: Reset column names to numeric indices first
                    df_cleaned.columns = range(len(df_cleaned.columns))
                    
                    # Step 2: DETECT TABLE BOUNDARIES with improved logic for better detection
                    table_boundaries = []
                    current_table = None
                    empty_row_count = 0
                    max_empty_gap = 3  # Allow up to 3 empty rows within a table
                    
                    for row_idx in range(len(df_cleaned)):
                        row = df_cleaned.iloc[row_idx]
                        row_has_data = False
                        first_col = None
                        last_col = None
                        
                        for col_idx in range(len(row)):
                            cell = row.iloc[col_idx]
                            if pd.notna(cell) and str(cell).strip():
                                row_has_data = True
                                if first_col is None:
                                    first_col = col_idx
                                last_col = col_idx
                        
                        if row_has_data:
                            empty_row_count = 0  # Reset empty row counter
                            
                            if current_table is None:
                                current_table = {
                                    'start_row': row_idx,
                                    'end_row': row_idx,
                                    'start_col': first_col,
                                    'end_col': last_col
                                }
                            else:
                                current_table['end_row'] = row_idx
                                current_table['start_col'] = min(current_table['start_col'], first_col)
                                current_table['end_col'] = max(current_table['end_col'], last_col)
                        else:
                            empty_row_count += 1
                            
                            # Only end the table if we have too many consecutive empty rows
                            if current_table is not None and empty_row_count > max_empty_gap:
                                table_boundaries.append(current_table)
                                current_table = None
                                empty_row_count = 0
                    
                    if current_table is not None:
                        table_boundaries.append(current_table)
                    
                    # Step 3: Smart cleaning - preserve table structure, clean outside tables
                    def is_inside_table(row_idx, col_idx):
                        """Check if cell is inside any detected table boundary"""
                        for table in table_boundaries:
                            if (table['start_row'] <= row_idx <= table['end_row'] and 
                                table['start_col'] <= col_idx <= table['end_col']):
                                return True
                        return False
                    
                    # Step 4: First remove completely empty columns outside tables
                    cols_to_keep = []
                    for col_idx in range(len(df_cleaned.columns)):
                        col_data = df_cleaned.iloc[:, col_idx]
                        
                        # Check if column intersects with any table
                        col_in_table = any(table['start_col'] <= col_idx <= table['end_col'] for table in table_boundaries)
                        
                        if col_in_table:
                            # Always keep columns that are part of tables
                            cols_to_keep.append(col_idx)
                        else:
                            # For columns outside tables, check if they have any data
                            has_data = col_data.notna().any() and any(str(cell).strip() != '' for cell in col_data if pd.notna(cell))
                            if has_data:
                                cols_to_keep.append(col_idx)
                    
                    # Create column mapping for updating table boundaries
                    old_to_new_col = {}
                    for new_idx, old_idx in enumerate(cols_to_keep):
                        old_to_new_col[old_idx] = new_idx
                    
                    if cols_to_keep:
                        df_cleaned = df_cleaned.iloc[:, cols_to_keep]
                        df_cleaned.columns = range(len(df_cleaned.columns))  # Reset to 0, 1, 2, ...
                        
                        # Update table boundaries after column removal
                        updated_table_boundaries = []
                        for table in table_boundaries:
                            # Map old column indices to new ones
                            new_start_col = None
                            new_end_col = None
                            
                            for old_col in range(table['start_col'], table['end_col'] + 1):
                                if old_col in old_to_new_col:
                                    new_col = old_to_new_col[old_col]
                                    if new_start_col is None:
                                        new_start_col = new_col
                                    new_end_col = new_col
                            
                            if new_start_col is not None and new_end_col is not None:
                                updated_table_boundaries.append({
                                    'start_row': table['start_row'],
                                    'end_row': table['end_row'],
                                    'start_col': new_start_col,
                                    'end_col': new_end_col
                                })
                        
                        table_boundaries = updated_table_boundaries
                    
                    # Step 5: Improved table-unit shifting with fallback for missed tables
                    # Process each table as a unit to maintain internal spacing and alignment
                    processed_tables = set()
                    
                    for idx in range(len(df_cleaned)):
                        row = df_cleaned.iloc[idx]
                        
                        # Skip completely empty rows
                        if row.isna().all() or all(str(cell).strip() == '' for cell in row if pd.notna(cell)):
                            continue
                        
                        # Check if this row is part of any table
                        current_table = None
                        for table in table_boundaries:
                            if table['start_row'] <= idx <= table['end_row']:
                                current_table = table
                                break
                        
                        if current_table and id(current_table) not in processed_tables:
                            # Process entire table as a unit
                            processed_tables.add(id(current_table))
                            
                            # Find the minimum leading empty columns across ALL data rows in this table
                            min_leading_empty = float('inf')
                            table_has_data = False
                            
                            for table_row_idx in range(current_table['start_row'], current_table['end_row'] + 1):
                                table_row = df_cleaned.iloc[table_row_idx]
                                
                                # Skip empty rows in the table
                                if table_row.isna().all() or all(str(cell).strip() == '' for cell in table_row if pd.notna(cell)):
                                    continue
                                
                                # Count leading empty cells in this row
                                leading_empty = 0
                                for cell in table_row:
                                    if pd.isna(cell) or str(cell).strip() == '':
                                        leading_empty += 1
                                    else:
                                        break
                                
                                # Only consider rows that have actual data
                                if leading_empty < len(table_row):
                                    min_leading_empty = min(min_leading_empty, leading_empty)
                                    table_has_data = True
                            
                            # If we found consistent leading empty space, shift the entire table
                            if table_has_data and min_leading_empty > 0 and min_leading_empty != float('inf'):
                                for table_row_idx in range(current_table['start_row'], current_table['end_row'] + 1):
                                    table_row = df_cleaned.iloc[table_row_idx]
                                    
                                    # Shift ALL rows in the table by the same amount (preserves internal structure)
                                    row_values = table_row.tolist()
                                    shifted = row_values[min_leading_empty:] + [None] * min_leading_empty
                                    df_cleaned.iloc[table_row_idx] = shifted
                        
                        elif not current_table:
                            # Rows OUTSIDE tables - shift individually
                            leading_empty = 0
                            for cell in row:
                                if pd.isna(cell) or str(cell).strip() == '':
                                    leading_empty += 1
                                else:
                                    break
                            
                            if leading_empty > 0 and leading_empty < len(row):
                                row_values = row.tolist()
                                shifted = row_values[leading_empty:] + [None] * leading_empty
                                df_cleaned.iloc[idx] = shifted
                    
                    # Step 5.5: Additional pass for any remaining rows with leading empty space
                    # This catches any rows that might have been missed by table detection
                    # BUT only shift rows that are OUTSIDE detected tables
                    for idx in range(len(df_cleaned)):
                        row = df_cleaned.iloc[idx]
                        
                        # Skip completely empty rows
                        if row.isna().all() or all(str(cell).strip() == '' for cell in row if pd.notna(cell)):
                            continue
                        
                        # Check if this row is inside any detected table
                        row_in_table = False
                        for table in table_boundaries:
                            if table['start_row'] <= idx <= table['end_row']:
                                row_in_table = True
                                break
                        
                        # Only shift if row is OUTSIDE tables
                        if not row_in_table:
                            # Count leading empty cells
                            leading_empty = 0
                            for cell in row:
                                if pd.isna(cell) or str(cell).strip() == '':
                                    leading_empty += 1
                                else:
                                    break
                            
                            # If row still has leading empty space, shift it
                            if leading_empty > 0 and leading_empty < len(row):
                                row_values = row.tolist()
                                shifted = row_values[leading_empty:] + [None] * leading_empty
                                df_cleaned.iloc[idx] = shifted
                    
                    # Step 6: Remove excessive empty rows (keep max 2 consecutive empty rows as separators)
                    rows_to_keep = []
                    consecutive_empty = 0
                    
                    for idx in range(len(df_cleaned)):
                        row = df_cleaned.iloc[idx]
                        is_empty = row.isna().all() or all(str(cell).strip() == '' for cell in row if pd.notna(cell))
                        
                        # Check if row contains header/footer text like "COMMERCIAL CONFIDENTIAL"
                        is_header_footer_row = False
                        for cell in row:
                            if pd.notna(cell):
                                cell_text = str(cell).strip().lower()
                                if 'commercial confidential' in cell_text:
                                    is_header_footer_row = True
                                    break
                        
                        if is_header_footer_row:
                            # Skip this row entirely
                            continue
                        elif is_empty:
                            consecutive_empty += 1
                            # Keep only first 2 consecutive empty rows
                            if consecutive_empty <= 2:
                                rows_to_keep.append(idx)
                        else:
                            consecutive_empty = 0
                            rows_to_keep.append(idx)
                    
                    # Apply the filter
                    if rows_to_keep:
                        df_cleaned = df_cleaned.iloc[rows_to_keep].reset_index(drop=True)
                    
                    # Write to Excel
                    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                        df_cleaned.to_excel(writer, sheet_name="Extracted Data", index=False, header=False)
                        
                        # Get workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets["Extracted Data"]
                        
                        # Define thick border style for tables
                        thick_border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium')
                        )
                        
                        # Detect table boundaries (non-empty cell regions)
                        table_ranges = []
                        current_table = None
                        
                        for row_idx in range(1, len(df_cleaned) + 1):
                            row_has_data = False
                            first_col = None
                            last_col = None
                            
                            for col_idx in range(1, len(df_cleaned.columns) + 1):
                                cell = worksheet.cell(row_idx, col_idx)
                                if cell.value is not None and str(cell.value).strip():
                                    row_has_data = True
                                    if first_col is None:
                                        first_col = col_idx
                                    last_col = col_idx
                            
                            if row_has_data:
                                if current_table is None:
                                    current_table = {
                                        'start_row': row_idx,
                                        'end_row': row_idx,
                                        'start_col': first_col,
                                        'end_col': last_col
                                    }
                                else:
                                    current_table['end_row'] = row_idx
                                    current_table['start_col'] = min(current_table['start_col'], first_col)
                                    current_table['end_col'] = max(current_table['end_col'], last_col)
                            else:
                                if current_table is not None:
                                    table_ranges.append(current_table)
                                    current_table = None
                        
                        if current_table is not None:
                            table_ranges.append(current_table)
                        
                        # Apply formatting to each table
                        for table_idx, table_range in enumerate(table_ranges):
                            start_row = table_range['start_row']
                            end_row = table_range['end_row']
                            start_col = table_range['start_col']
                            end_col = table_range['end_col']
                            
                            # Apply thick borders and formatting to table cells
                            for row_idx in range(start_row, end_row + 1):
                                for col_idx in range(start_col, end_col + 1):
                                    cell = worksheet.cell(row_idx, col_idx)
                                    
                                    # Apply thick border only to table cells
                                    cell.border = thick_border
                                    
                                    # Center align
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)
                            for cell in column:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Create Risk Grades sheet if we have risk grade for this file
                        risk_grade = self.file_risk_grades.get(file_name, None)
                        if risk_grade:
                            risk_sheet = workbook.create_sheet("Risk Grades")
                            
                            # Header
                            risk_sheet['A1'] = "File Name"
                            risk_sheet['B1'] = "Risk Grade"
                            
                            # Data
                            risk_sheet['A2'] = file_name
                            risk_sheet['B2'] = risk_grade
                            
                            # Format header
                            for col in ['A', 'B']:
                                cell = risk_sheet[f'{col}1']
                                cell.font = Font(bold=True, size=12)
                                cell.fill = PatternFill(start_color="1f538d", end_color="1f538d", fill_type="solid")
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.border = thick_border
                            
                            # Format risk grade cell
                            risk_sheet['B2'].font = Font(bold=True, size=14, color="FF0000")
                            risk_sheet['B2'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            risk_sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')
                            risk_sheet['B2'].border = thick_border
                            
                            # Adjust column widths
                            risk_sheet.column_dimensions['A'].width = 50
                            risk_sheet.column_dimensions['B'].width = 15
                
                self.status_label.configure(text=f"✓ Exported {len(self.file_previews)} files")
                messagebox.showinfo("Success", f"Exported {len(self.file_previews)} files to:\n{save_folder}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data:\n{str(e)}")
            self.status_label.configure(text="✗ Export failed")
            import traceback
            traceback.print_exc()
    
    def update_database_file_list(self):
        """Enable/disable build button based on available data"""
        try:
            # Enable build button if there is at least one preview
            if any((not df.empty) for df in self.file_previews.values()):
                self.db_build_btn.configure(state="normal")
            else:
                self.db_build_btn.configure(state="disabled")
            self.db_export_btn.configure(state="disabled")
        except Exception as e:
            print(f"update_database_file_list error: {e}")
    
    def build_database(self):
        """Build database from PARTICULARS, SUMMARY CREDIT INFORMATION, CREDIT SCORE, SHAREHOLDING INTEREST, CCRIS ENTITY, SUBJECT STATUS, and KEY STATISTICS tables"""
        self.show_loading_screen("Building database from multiple tables...")
        
        def build_db_thread():
            try:
                # Include all files from file_previews
                all_files = list(self.file_previews.keys())
                
                if not all_files:
                    def show_info():
                        self.hide_loading_screen()
                        messagebox.showinfo("Info", "No files available to build database.")
                    self.after(0, show_info)
                    return
                
                print(f"🏗️  Building 7-table database from {len(all_files)} files...")
                database_rows = []
                successful_extractions = 0
                failed_extractions = 0
                total_files = len(all_files)
                
                for file_idx, file_name in enumerate(all_files):
                    progress = int((file_idx / total_files) * 100)
                    self.after(0, lambda p=progress: self.update_loading_progress(p))
                    
                    print(f"\n📄 Processing file {file_idx + 1}/{total_files}: {file_name}")
                    df = self.file_previews[file_name]
                    if df.empty:
                        failed_extractions += 1
                        continue
                    
                    # Extract data from this file's DataFrame (can return multiple rows)
                    extracted_rows = self.extract_database_row(df, file_name)
                    
                    if extracted_rows and isinstance(extracted_rows, list):
                        # Multiple rows returned (multiple contributing factors)
                        valid_rows = [row for row in extracted_rows if row and any(row.values())]
                        if valid_rows:
                            database_rows.extend(valid_rows)
                            successful_extractions += 1
                            # Debug: Show sample of extracted data
                            sample_row = valid_rows[0]
                            print(f"   📋 Sample data: i-SCORE={sample_row.get('i_SCORE', 'N/A')}, Risk_Grade={sample_row.get('Risk_Grade', 'N/A')}")
                            print(f"   📋 Contributing Factor: {sample_row.get('Key_Contributing_Factor', 'N/A')[:50]}...")
                        else:
                            failed_extractions += 1
                    else:
                        failed_extractions += 1
                
                # Create database DataFrame
                if database_rows:
                    self.database_df = pd.DataFrame(database_rows)
                else:
                    self.database_df = pd.DataFrame()
                
                def finish_build():
                    self.update_loading_progress(100)
                    time.sleep(0.2)
                    self.hide_loading_screen()
                    
                    # Update preview
                    self.update_database_preview()
                    
                    # Enable export button if we have data
                    if not self.database_df.empty:
                        self.db_export_btn.configure(state="normal")
                    
                    summary_msg = f"Multi-Table Database Build Complete!\n\n"
                    summary_msg += f"✅ Successful extractions: {successful_extractions}/{total_files} files\n"
                    summary_msg += f"❌ Failed extractions: {failed_extractions}/{total_files} files\n"
                    summary_msg += f"📊 Total database records: {len(database_rows)}\n\n"
                    summary_msg += f"Tables extracted:\n"
                    summary_msg += f"• PARTICULARS OF THE SUBJECT PROVIDED BY YOU (5 fields)\n"
                    summary_msg += f"• SUMMARY CREDIT INFORMATION (10 fields)\n"
                    summary_msg += f"• CREDIT SCORE (3 fields with Risk Grade calculation)\n"
                    summary_msg += f"• SHAREHOLDING INTEREST (9 fields)\n"
                    summary_msg += f"• CCRIS ENTITY SELECTED BY YOU (1 field)\n"
                    summary_msg += f"• SUBJECT STATUS (1 field)\n"
                    summary_msg += f"• KEY STATISTICS (4 fields - split merged cells)\n\n"
                    summary_msg += f"Note: Creates multiple rows for each combination of:\n"
                    summary_msg += f"- Contributing factors (bullet-point separated)\n"
                    summary_msg += f"- Business interests (from shareholding table)\n"
                    summary_msg += f"- Facility records (earliest + latest 3 approved)\n"
                    summary_msg += f"Total fields: 35 (5 + 10 + 3 + 9 + 1 + 1 + 4 + 2 tracking fields)"
                    
                    self.status_label.configure(text=f"✓ 7-Table Database: {successful_extractions}/{total_files} successful")
                    messagebox.showinfo("Multi-Table Database Complete", summary_msg)
                
                self.after(0, finish_build)
                
            except Exception as e:
                def show_error():
                    self.hide_loading_screen()
                    messagebox.showerror("Error", f"Failed to build database:\n{str(e)}")
                    import traceback
                    traceback.print_exc()
                self.after(0, show_error)
        
        # Start thread
        thread = threading.Thread(target=build_db_thread, daemon=True)
        thread.start()
    
    def extract_database_row(self, df, file_name):
        """Extract data from PARTICULARS, SUMMARY CREDIT INFORMATION, CREDIT SCORE, SHAREHOLDING INTEREST, CCRIS ENTITY, SUBJECT STATUS, and KEY STATISTICS tables"""
        try:
            base_row_data = {}
            
            
            # Helper function to find value in PARTICULARS table
            def find_particulars_value(label_text):
                """Find value from PARTICULARS OF THE SUBJECT PROVIDED BY YOU table"""
                particulars_found = False
                
                for idx, row in df.iterrows():
                    # Check if we found the PARTICULARS table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "PARTICULARS OF THE SUBJECT PROVIDED BY YOU" in cell_str.upper():
                            particulars_found = True
                            break
                    
                    if particulars_found:
                        # Look for the specific label in the rows following the header
                        for search_idx in range(idx + 1, min(idx + 10, len(df))):  # Search next 10 rows
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            for col_idx, cell in enumerate(search_row):
                                cell_str = str(cell).strip() if pd.notna(cell) else ""
                                if label_text.lower() in cell_str.lower():
                                    # Found the label, get value from next column
                                    if col_idx + 1 < len(search_row):
                                        value = search_row.iloc[col_idx + 1]
                                        result = str(value).strip() if pd.notna(value) else ""
                                        return result
                        break
                
                return ""
            
            # Helper function to find value in SUMMARY CREDIT INFORMATION table
            def find_summary_credit_value(label_text):
                """Find value from SUMMARY CREDIT INFORMATION table"""
                summary_found = False
                
                for idx, row in df.iterrows():
                    # Check if we found the SUMMARY CREDIT INFORMATION table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "SUMMARY CREDIT INFORMATION" in cell_str.upper():
                            summary_found = True
                            break
                    
                    if summary_found:
                        # Look for the specific label in the rows following the header
                        for search_idx in range(idx + 1, min(idx + 15, len(df))):  # Search next 15 rows
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            for col_idx, cell in enumerate(search_row):
                                cell_str = str(cell).strip() if pd.notna(cell) else ""
                                if label_text.lower() in cell_str.lower():
                                    # Found the label, get value from next column
                                    if col_idx + 1 < len(search_row):
                                        value = search_row.iloc[col_idx + 1]
                                        result = str(value).strip() if pd.notna(value) else ""
                                        return result
                        break
                
                return ""
            
            # Helper function to find value in CREDIT SCORE table
            def find_credit_score_value(label_text):
                """Find value from CREDIT SCORE table"""
                credit_score_found = False
                
                for idx, row in df.iterrows():
                    # Check if we found the CREDIT SCORE table header - more flexible matching
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "CREDIT SCORE" in cell_str.upper() or "i-SCORE" in cell_str.upper():
                            credit_score_found = True
                            break
                    
                    if credit_score_found:
                        # Look for the specific label in the rows following the header
                        for search_idx in range(idx + 1, min(idx + 20, len(df))):  # Search next 20 rows
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            for col_idx, cell in enumerate(search_row):
                                cell_str = str(cell).strip() if pd.notna(cell) else ""
                                
                                # Special handling for i-SCORE - look for numbers
                                if "i-score" in label_text.lower():
                                    if "i-score" in cell_str.lower() or re.search(r'\bi-?score\b', cell_str, re.IGNORECASE):
                                        # Found i-SCORE label, get value from next column or same cell
                                        if col_idx + 1 < len(search_row):
                                            value = search_row.iloc[col_idx + 1]
                                        else:
                                            # Try to extract number from same cell
                                            numbers = re.findall(r'\d+', cell_str)
                                            value = numbers[0] if numbers else ""
                                        result = str(value).strip() if pd.notna(value) else ""
                                        return result
                                
                                # Special handling for Key Contributing Factors - look for bullet points
                                elif "contributing" in label_text.lower():
                                    if "contributing" in cell_str.lower() or "factor" in cell_str.lower():
                                        # Found contributing factors, collect all text from this and following rows
                                        factors_text = ""
                                        # Check current cell first
                                        if col_idx + 1 < len(search_row):
                                            next_cell = search_row.iloc[col_idx + 1]
                                            if pd.notna(next_cell):
                                                factors_text += str(next_cell).strip()
                                        
                                        # Define section headers that indicate end of contributing factors
                                        section_headers = [
                                            "SHAREHOLDING INTEREST", "INTEREST IN COMPANY", 
                                            "SUMMARY CREDIT INFORMATION", "KEY STATISTICS",
                                            "PARTICULARS", "CREDIT REPORT", "NOTE:"
                                        ]
                                        
                                        # Check following rows for more factors
                                        for factor_idx in range(search_idx + 1, min(search_idx + 10, len(df))):
                                            if factor_idx >= len(df):
                                                break
                                            factor_row = df.iloc[factor_idx]
                                            
                                            # Check if we've hit a new section header
                                            row_text = " ".join([str(cell).strip() for cell in factor_row if pd.notna(cell)])
                                            if any(header in row_text.upper() for header in section_headers):
                                                break  # Stop collecting factors
                                            
                                            for factor_col_idx, factor_cell in enumerate(factor_row):
                                                factor_str = str(factor_cell).strip() if pd.notna(factor_cell) else ""
                                                # Look for bullet points or continuation text
                                                if "•" in factor_str or (len(factor_str) > 10 and not any(keyword in factor_str.lower() for keyword in ["summary", "particular", "experian", "page", "note:", "shareholding"])):
                                                    if factors_text:
                                                        factors_text += " " + factor_str
                                                    else:
                                                        factors_text = factor_str
                                        
                                        result = factors_text.strip()
                                        return result
                                
                                # Regular label matching
                                elif label_text.lower() in cell_str.lower():
                                    # Found the label, get value from next column
                                    if col_idx + 1 < len(search_row):
                                        value = search_row.iloc[col_idx + 1]
                                        result = str(value).strip() if pd.notna(value) else ""
                                        return result
                        break
                
                return ""
            
            # Helper function to extract KEY STATISTICS Earliest/Latest Approved Facilities
            def find_key_statistics_facilities():
                """Find Earliest and Latest 3 Approved Facilities from KEY STATISTICS table"""
                key_stats_found = False
                earliest_facility = {"type": "-", "date": "-"}
                latest_facilities = []
                
                for idx, row in df.iterrows():
                    # Check if we found the KEY STATISTICS table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "KEY STATISTICS" in cell_str.upper():
                            key_stats_found = True
                            break
                    
                    if key_stats_found:
                        # Look for Earliest and Latest facilities in the following rows
                        for search_idx in range(idx + 1, min(idx + 20, len(df))):
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            row_data = [str(cell).strip() if pd.notna(cell) else "" for cell in search_row]
                            
                            # Check for "Earliest Approved Facilites"
                            for col_idx, cell_str in enumerate(row_data):
                                if "Earliest Approved" in cell_str and "Facility Type" in cell_str:
                                    # Found Earliest row, get facility type and date from next columns
                                    if col_idx + 1 < len(row_data):
                                        earliest_facility["type"] = row_data[col_idx + 1].strip()
                                    if col_idx + 2 < len(row_data):
                                        earliest_facility["date"] = row_data[col_idx + 2].strip()
                                    break
                                
                                # Check for "Latest 3 Approved Facilites"
                                elif "Latest 3 Approved" in cell_str or "Latest Approved" in cell_str:
                                    # Found Latest 3 header row, get first facility from same row
                                    if col_idx + 1 < len(row_data) and col_idx + 2 < len(row_data):
                                        facility_type = row_data[col_idx + 1].strip()
                                        facility_date = row_data[col_idx + 2].strip()
                                        if facility_type and facility_type != "":
                                            latest_facilities.append({"type": facility_type, "date": facility_date})
                                    
                                    # Continue looking for the remaining 2 facilities in next rows
                                    for next_idx in range(search_idx + 1, min(search_idx + 5, len(df))):
                                        if next_idx >= len(df):
                                            break
                                        next_row = df.iloc[next_idx]
                                        next_row_data = [str(cell).strip() if pd.notna(cell) else "" for cell in next_row]
                                        
                                        # Look for rows with facility type and date (should be in first few columns after blank)
                                        # Structure: [blank] | FACILITY TYPE | DATE
                                        for nc_idx in range(len(next_row_data) - 1):
                                            if next_row_data[nc_idx] and next_row_data[nc_idx + 1]:
                                                # Check if this looks like a facility type (all caps, has words)
                                                if len(next_row_data[nc_idx]) > 5 and next_row_data[nc_idx].isupper():
                                                    facility_type = next_row_data[nc_idx].strip()
                                                    facility_date = next_row_data[nc_idx + 1].strip()
                                                    
                                                    # Validate date format (DD-MM-YYYY)
                                                    if re.match(r'\d{2}-\d{2}-\d{4}', facility_date):
                                                        latest_facilities.append({"type": facility_type, "date": facility_date})
                                                        break
                                        
                                        # Stop if we have 3 latest facilities
                                        if len(latest_facilities) >= 3:
                                            break
                                    break
                        
                        # Stop after processing KEY STATISTICS section
                        if earliest_facility["type"] != "-" or len(latest_facilities) > 0:
                            break
                
                # Ensure we have exactly 3 latest facilities (pad with empty if needed)
                while len(latest_facilities) < 3:
                    latest_facilities.append({"type": "-", "date": "-"})
                
                # Only take first 3 if more were found
                latest_facilities = latest_facilities[:3]
                
                if not key_stats_found:
                    pass
                
                return earliest_facility, latest_facilities
            
            # Helper function to extract SHAREHOLDING INTEREST data
            def find_shareholding_interests():
                """Find all business interests from SHAREHOLDING INTEREST table"""
                interests = []
                shareholding_found = False
                header_found = False
                
                for idx, row in df.iterrows():
                    # Check if we found the SHAREHOLDING INTEREST table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "SHAREHOLDING INTEREST" in cell_str.upper() or ("INTEREST IN COMPANY" in cell_str.upper() and "BUSINESS" in cell_str.upper()):
                            shareholding_found = True
                            break
                    
                    if shareholding_found and not header_found:
                        # Look for the column header row (No, Name, Position, etc.)
                        for header_idx in range(idx + 1, min(idx + 10, len(df))):
                            if header_idx >= len(df):
                                break
                            header_row = df.iloc[header_idx]
                            header_text = "".join([str(cell).strip() if pd.notna(cell) else "" for cell in header_row])
                            
                            # Check if this row contains the column headers
                            if ("No" in header_text and "Name" in header_text and "Position" in header_text and 
                                "Appointed" in header_text):
                                header_found = True
                                
                                # Now look for data rows after the header
                                for data_idx in range(header_idx + 1, min(header_idx + 50, len(df))):
                                    if data_idx >= len(df):
                                        break
                                    data_row = df.iloc[data_idx]
                                    
                                    # Extract row data
                                    row_data = []
                                    for col_idx, cell in enumerate(data_row):
                                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                                        row_data.append(cell_str)
                                    
                                    # Check if this is a valid data row (starts with number)
                                    if (len(row_data) > 0 and row_data[0].isdigit() and
                                        len([x for x in row_data if x]) >= 4):  # At least 4 non-empty columns
                                        
                                        # Parse according to actual table structure:
                                        # No | Name | Position | Appointed | Business Expiry Date | Shareholding | % | Remark | Last Updated by Experian
                                        interest_data = {
                                            'No': row_data[0] if len(row_data) > 0 else "-",
                                            'Name': row_data[1] if len(row_data) > 1 else "-",
                                            'Position': row_data[2] if len(row_data) > 2 else "-",
                                            'Appointed': row_data[3] if len(row_data) > 3 else "-",
                                            'Business_Expiry_Date': row_data[4] if len(row_data) > 4 else "-",
                                            'Shareholding': row_data[5] if len(row_data) > 5 else "-",
                                            'Percentage': row_data[6] if len(row_data) > 6 else "-",
                                            'Remark': row_data[7] if len(row_data) > 7 else "-",
                                            'Last_Updated_by_Experian': row_data[8] if len(row_data) > 8 else "-"
                                        }
                                        
                                        # Clean and format data - replace empty with "-"
                                        for key, value in interest_data.items():
                                            if not value or value.strip() == "" or pd.isna(value):
                                                interest_data[key] = "-"
                                            else:
                                                # Clean extra spaces and newlines
                                                clean_value = re.sub(r'\s+', ' ', str(value).strip())
                                                interest_data[key] = clean_value
                                        
                                        interests.append(interest_data)
                                break  # Found headers, processed data
                        break  # Found shareholding table
                
                if not interests:
                    # Return empty interest to maintain structure
                    interests.append({
                        'No': "-",
                        'Name': "-",
                        'Position': "-",
                        'Appointed': "-",
                        'Business_Expiry_Date': "-",
                        'Shareholding': "-",
                        'Percentage': "-",
                        'Remark': "-",
                        'Last_Updated_by_Experian': "-"
                    })
                
                return interests
            
            # Helper function to extract CCRIS ENTITY data
            def find_ccris_entity_key():
                """Find CCRIS Entity Key from CCRIS ENTITY SELECTED BY YOU table"""
                ccris_found = False
                
                for idx, row in df.iterrows():
                    # Check if we found the CCRIS ENTITY table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "CCRIS ENTITY SELECTED BY YOU" in cell_str.upper():
                            ccris_found = True
                            break
                    
                    if ccris_found:
                        # Look for CCRIS Entity Key in the following rows
                        for search_idx in range(idx + 1, min(idx + 20, len(df))):
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            
                            for col_idx, cell in enumerate(search_row):
                                cell_str = str(cell).strip() if pd.notna(cell) else ""
                                if "CCRIS Entity Key" in cell_str:
                                    # Found the label, get value from next column
                                    if col_idx + 1 < len(search_row):
                                        value = search_row.iloc[col_idx + 1]
                                        result = str(value).strip() if pd.notna(value) else "-"
                                        return result if result else "-"
                        break
                
                return "-"
            
            # Helper function to extract SUMMARY CREDIT REPORT data
            def find_summary_credit_data():
                """Find data from SUMMARY CREDIT REPORT table with custom field naming"""
                summary_found = False
                result = {
                    'A_App_No_Application': "-",
                    'A_App_Ttl_Amnt': "-", 
                    'B_Pend_No_Application': "-",
                    'B_Pend_Ttl_Amnt': "-"
                }
                
                for idx, row in df.iterrows():
                    # Check if we found the SUMMARY CREDIT REPORT table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "SUMMARY CREDIT REPORT" in cell_str.upper():
                            summary_found = True
                            break
                    
                    if summary_found:
                        # Look for the specific rows in the following rows
                        for search_idx in range(idx + 1, min(idx + 20, len(df))):
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            
                            # Convert row to list for easier processing
                            row_data = [str(cell).strip() if pd.notna(cell) else "" for cell in search_row]
                            
                            # Check for "A. Approved for past 12 months" row
                            for col_idx, cell_str in enumerate(row_data):
                                if "A. Approved for past 12 months" in cell_str:
                                    # Found A row, extract No. of Applications and Total Amount
                                    if col_idx + 1 < len(row_data):
                                        result['A_App_No_Application'] = row_data[col_idx + 1].strip()
                                    if col_idx + 2 < len(row_data):
                                        result['A_App_Ttl_Amnt'] = row_data[col_idx + 2].strip()
                                    break
                                    
                                elif "B. Pending" in cell_str:
                                    # Found B row, extract No. of Applications and Total Amount  
                                    if col_idx + 1 < len(row_data):
                                        result['B_Pend_No_Application'] = row_data[col_idx + 1].strip()
                                    if col_idx + 2 < len(row_data):
                                        result['B_Pend_Ttl_Amnt'] = row_data[col_idx + 2].strip()
                                    break
                        break
                
                # Clean results - replace empty with "-"
                for key, value in result.items():
                    if not value or value == "":
                        result[key] = "-"
                
                return result

            # Helper function to extract Subject Status data
            def find_warning_remark():
                """Find Warning Remark from Subject Status table"""
                subject_status_found = False
                
                for idx, row in df.iterrows():
                    # Check if we found the Subject Status table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "Subject Status" in cell_str:
                            subject_status_found = True
                            break
                    
                    if subject_status_found:
                        # Look for Warning Remark in the following rows
                        for search_idx in range(idx + 1, min(idx + 10, len(df))):
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            
                            for col_idx, cell in enumerate(search_row):
                                cell_str = str(cell).strip() if pd.notna(cell) else ""
                                if "Warning Remark" in cell_str:
                                    # Found the label, get value from next column
                                    if col_idx + 1 < len(search_row):
                                        value = search_row.iloc[col_idx + 1]
                                        result = str(value).strip() if pd.notna(value) else "-"
                                        return result if result else "-"
                        break
                
                return "-"
            
            # Helper function to extract SUMMARY OF POTENTIAL & CURRENT LIABILITIES data
            def find_potential_liabilities_data():
                """Find data from SUMMARY OF POTENTIAL & CURRENT LIABILITIES table with hybrid structure"""
                liabilities_found = False
                result = {
                    'AsBorr_Outstanding_RM': "-",
                    'AsBorr_Total_Limit_RM': "-",
                    'AsBorr_FEC_Limit_RM': "-",
                    'Legal_Action_Taken': "-",
                    'Special_Attention_Account': "-"
                }
                
                for idx, row in df.iterrows():
                    # Check if we found the SUMMARY OF POTENTIAL & CURRENT LIABILITIES table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "SUMMARY OF POTENTIAL" in cell_str.upper() and "CURRENT LIABILITIES" in cell_str.upper():
                            liabilities_found = True
                            break
                    
                    if liabilities_found:
                        # Look for the specific rows in the following rows
                        for search_idx in range(idx + 1, min(idx + 30, len(df))):
                            if search_idx >= len(df):
                                break
                            search_row = df.iloc[search_idx]
                            
                            # Convert row to list for easier processing
                            row_data = [str(cell).strip() if pd.notna(cell) else "" for cell in search_row]
                            
                            # Check for "As Borrower" row (row-based data with 3 values)
                            for col_idx, cell_str in enumerate(row_data):
                                if "As Borrower" in cell_str:
                                    # Found As Borrower row, extract the 3 numeric values
                                    if col_idx + 1 < len(row_data):
                                        result['AsBorr_Outstanding_RM'] = row_data[col_idx + 1].strip()
                                    if col_idx + 2 < len(row_data):
                                        result['AsBorr_Total_Limit_RM'] = row_data[col_idx + 2].strip()
                                    if col_idx + 3 < len(row_data):
                                        result['AsBorr_FEC_Limit_RM'] = row_data[col_idx + 3].strip()
                                    break
                                    
                                elif "Legal Action Taken" in cell_str:
                                    # Found Legal Action Taken, extract value from next column
                                    if col_idx + 1 < len(row_data):
                                        result['Legal_Action_Taken'] = row_data[col_idx + 1].strip()
                                    break
                                    
                                elif "Special Attention Account" in cell_str:
                                    # Found Special Attention Account, extract value from next column
                                    if col_idx + 1 < len(row_data):
                                        result['Special_Attention_Account'] = row_data[col_idx + 1].strip()
                                    break
                        break
                
                # Clean results - replace empty with "-"
                for key, value in result.items():
                    if not value or value == "":
                        result[key] = "-"
                
                return result
            
            # Helper function to extract Legal Suits and Bankruptcy data
            def find_legal_suits_bankruptcy_data():
                """Find data from Legal Suits and Bankruptcy tables where totals are in headers"""
                result = {
                    'Legal_Suits_Defendant': "-",
                    'Legal_Suits_Plaintiff': "-",
                    'Bankruptcy_Action': "-"
                }
                
                for idx, row in df.iterrows():
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        
                        # Check for "LEGAL SUITS - SUBJECT AS DEFENDANT Total: X"
                        if "LEGAL SUITS" in cell_str.upper() and "SUBJECT AS DEFENDANT" in cell_str.upper() and "TOTAL" in cell_str.upper():
                            # Extract number from "Total: 0" pattern
                            match = re.search(r'Total:\s*(\d+)', cell_str, re.IGNORECASE)
                            if match:
                                result['Legal_Suits_Defendant'] = match.group(1)
                        
                        # Check for "LEGAL SUITS - SUBJECT AS PLAINTIFF Total: X"
                        elif "LEGAL SUITS" in cell_str.upper() and "SUBJECT AS PLAINTIFF" in cell_str.upper() and "TOTAL" in cell_str.upper():
                            # Extract number from "Total: 0" pattern
                            match = re.search(r'Total:\s*(\d+)', cell_str, re.IGNORECASE)
                            if match:
                                result['Legal_Suits_Plaintiff'] = match.group(1)
                        
                        # Check for "BANKRUPTCY ACTION" header
                        elif "BANKRUPTCY ACTION" in cell_str.upper():
                            # Look for "Total: X" in the following rows
                            for search_idx in range(idx + 1, min(idx + 5, len(df))):
                                if search_idx >= len(df):
                                    break
                                search_row = df.iloc[search_idx]
                                
                                for search_col_idx, search_cell in enumerate(search_row):
                                    search_str = str(search_cell).strip() if pd.notna(search_cell) else ""
                                    if "Total:" in search_str or "Total :" in search_str:
                                        # Extract number from "Total: 0" pattern
                                        match = re.search(r'Total:\s*(\d+)', search_str, re.IGNORECASE)
                                        if match:
                                            result['Bankruptcy_Action'] = match.group(1)
                                            break
                                
                                if result['Bankruptcy_Action'] != "-":
                                    break
                
                return result
            
            # Helper function to extract KEY STATISTICS data
            def find_key_statistics_data():
                """Find data from KEY STATISTICS table"""
                key_stats_found = False
                key_stats_start_idx = -1
                result = {
                    'SF_No_of_Facilities': "-",
                    'SF_Total_Outstanding_Balance_RM': "-",
                    'SF_Total_Outstanding_Balance_Against_Total_Limit': "-",
                    'SF_Highest_No_of_Installments_Arrears_Last_12_months': "-",
                    'UF_No_of_Facilities': "-",
                    'UF_Total_Outstanding_Balance_RM': "-",
                    'UF_Total_Outstanding_Balance_Against_Total_Limit': "-",
                    'UF_Highest_No_of_Installments_Arrears_Last_12_months': "-",
                    'CC_Average_Utilisation_Last_6_months': "-",
                    'ORC_Average_Utilisation_Last_6_months': "-",
                    'CHC_Min_Utilisation_Last_12_months_RM': "-",
                    'CHC_Max_Utilisation_Last_12_months_RM': "-",
                    'NHEF_No_of_Accounts': "-",
                    'LL_No_of_Accounts': "-",
                    'FL_No_of_Accounts': "-"
                }
                
                # Track which section we're in
                current_section = None
                
                for idx, row in df.iterrows():
                    # Check if we found the KEY STATISTICS table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "KEY STATISTICS" in cell_str.upper():
                            key_stats_found = True
                            key_stats_start_idx = idx
                            break
                    
                    if key_stats_found:
                        row_data = [str(cell).strip() if pd.notna(cell) else "" for cell in row]
                        
                        # Identify section headers
                        for col_idx, cell_str in enumerate(row_data):
                            if "Secured Facilities" in cell_str:
                                current_section = "SF"
                            elif "Unsecured Facilities" in cell_str:
                                current_section = "UF"
                            elif cell_str == "Credit Card":
                                current_section = "CC"
                            elif "Other Revolving Credits" in cell_str:
                                current_section = "ORC"
                            elif "Charge Card" in cell_str:
                                current_section = "CHC"
                            elif "National Higher Educational Financing" in cell_str:
                                current_section = "NHEF"
                            elif "Local Lenders" in cell_str:
                                current_section = "LL"
                            elif "Foreign Lenders" in cell_str:
                                current_section = "FL"
                            
                            # Extract data based on current section
                            if current_section == "SF":
                                if "No. of Facilities" in cell_str and col_idx + 1 < len(row_data):
                                    result['SF_No_of_Facilities'] = row_data[col_idx + 1]
                                elif "Total Outstanding Balance (RM)" in cell_str and col_idx + 1 < len(row_data):
                                    result['SF_Total_Outstanding_Balance_RM'] = row_data[col_idx + 1]
                                elif "Total Outstanding Balance Against Total Limit" in cell_str and col_idx + 1 < len(row_data):
                                    result['SF_Total_Outstanding_Balance_Against_Total_Limit'] = row_data[col_idx + 1]
                                elif "Highest No. of Installments Arrears Last 12 months" in cell_str and col_idx + 1 < len(row_data):
                                    result['SF_Highest_No_of_Installments_Arrears_Last_12_months'] = row_data[col_idx + 1]
                            
                            elif current_section == "UF":
                                if "No. of Facilities" in cell_str and col_idx + 1 < len(row_data):
                                    result['UF_No_of_Facilities'] = row_data[col_idx + 1]
                                elif "Total Outstanding Balance (RM)" in cell_str and col_idx + 1 < len(row_data):
                                    result['UF_Total_Outstanding_Balance_RM'] = row_data[col_idx + 1]
                                elif "Total Outstanding Balance Against Total Limit" in cell_str and col_idx + 1 < len(row_data):
                                    result['UF_Total_Outstanding_Balance_Against_Total_Limit'] = row_data[col_idx + 1]
                                elif "Highest No. of Installments Arrears Last 12 months" in cell_str and col_idx + 1 < len(row_data):
                                    result['UF_Highest_No_of_Installments_Arrears_Last_12_months'] = row_data[col_idx + 1]
                            
                            elif current_section == "CC":
                                if "Average Utilisation Last 6 months" in cell_str and col_idx + 1 < len(row_data):
                                    result['CC_Average_Utilisation_Last_6_months'] = row_data[col_idx + 1]
                            
                            elif current_section == "ORC":
                                if "Average Utilisation Last 6 months" in cell_str and col_idx + 1 < len(row_data):
                                    result['ORC_Average_Utilisation_Last_6_months'] = row_data[col_idx + 1]
                            
                            elif current_section == "CHC":
                                if "Min Utilisation Last 12 months (RM)" in cell_str and col_idx + 1 < len(row_data):
                                    result['CHC_Min_Utilisation_Last_12_months_RM'] = row_data[col_idx + 1]
                                elif "Max Utilisation Last 12 months (RM)" in cell_str and col_idx + 1 < len(row_data):
                                    result['CHC_Max_Utilisation_Last_12_months_RM'] = row_data[col_idx + 1]
                            
                            elif current_section == "NHEF":
                                if "No. of Accounts" in cell_str and col_idx + 1 < len(row_data):
                                    result['NHEF_No_of_Accounts'] = row_data[col_idx + 1]
                            
                            elif current_section == "LL":
                                if "No. of Accounts" in cell_str and col_idx + 1 < len(row_data):
                                    result['LL_No_of_Accounts'] = row_data[col_idx + 1]
                            
                            elif current_section == "FL":
                                if "No. of Accounts" in cell_str and col_idx + 1 < len(row_data):
                                    result['FL_No_of_Accounts'] = row_data[col_idx + 1]
                        
                        # Stop after sufficient rows (typically within 40-50 rows from start)
                        if key_stats_start_idx != -1 and idx > key_stats_start_idx + 50:
                            break
                
                if not key_stats_found:
                    pass
                
                # Clean results - replace empty with "-"
                for key, value in result.items():
                    if not value or value == "":
                        result[key] = "-"
                
                return result
            
            # Helper function to extract TRADE / CREDIT REFERENCE data (can have multiple records)
            def find_trade_credit_reference_data():
                """Find all data from TRADE / CREDIT REFERENCE (CR) table - returns list of records"""
                trade_found = False
                trade_start_idx = -1
                trade_records = []  # Store multiple trade/credit reference records
                current_record = {}
                
                # Mapping of label patterns to result keys
                label_mapping = {
                    'Subject Name': 'TCR_Subject_Name',
                    'Creditor\'s Name': 'TCR_Creditors_Name',
                    'Creditor\'s Contact': 'TCR_Creditors_Contact',
                    'Ref No': 'TCR_Ref_No',
                    'Industry': 'TCR_Industry',
                    'Solicitor\'s Name': 'TCR_Solicitors_Name',
                    'Guarantor / Owner': 'TCR_Guarantor_Owner',
                    'Subject ID': 'TCR_Subject_ID',
                    'Amount Due': 'TCR_Amount_Due',
                    'Aging Days': 'TCR_Aging_Days',
                    'Debt Type': 'TCR_Debt_Type',
                    'Document/Status Date': 'TCR_Document_Status_Date',
                    'Solicitor\'s Contact': 'TCR_Solicitors_Contact',
                    'Remark': 'TCR_Remark'
                }
                
                for idx, row in df.iterrows():
                    # Check if we found the TRADE / CREDIT REFERENCE table header
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        if "TRADE / CREDIT REFERENCE" in cell_str.upper() and "(CR)" in cell_str.upper():
                            trade_found = True
                            trade_start_idx = idx
                            break
                    
                    if trade_found:
                        # Extract label-value pairs from the table
                        row_data = [str(cell).strip() if pd.notna(cell) else "" for cell in row]
                        
                        # Check if this row starts a new record (contains 'Creditor's Name')
                        has_creditor_name = False
                        for col_idx, cell_str in enumerate(row_data):
                            if cell_str == "Creditor's Name" and col_idx + 1 < len(row_data):
                                creditor_value = row_data[col_idx + 1].strip()
                                if creditor_value and creditor_value != "":
                                    # Save previous record if it exists
                                    if current_record and current_record.get('TCR_Creditors_Name', '-') != '-':
                                        trade_records.append(current_record)
                                    # Start new record
                                    current_record = {key: "-" for key in label_mapping.values()}
                                    has_creditor_name = True
                                    break
                        
                        # Extract all fields in current row
                        for col_idx, cell_str in enumerate(row_data):
                            for label_pattern, result_key in label_mapping.items():
                                if cell_str == label_pattern:
                                    if col_idx + 1 < len(row_data):
                                        value = row_data[col_idx + 1].strip()
                                        if value and value != "":
                                            if current_record or has_creditor_name:
                                                if not current_record:
                                                    current_record = {key: "-" for key in label_mapping.values()}
                                                current_record[result_key] = value
                                    break
                        
                        # Stop after finding enough data rows
                        if trade_start_idx != -1 and idx > trade_start_idx + 30:
                            break
                
                # Add last record if exists
                if current_record and current_record.get('TCR_Creditors_Name', '-') != '-':
                    trade_records.append(current_record)
                
                if not trade_found:
                    pass
                
                # If no records found, return empty list (will use default "-" in row generation)
                if not trade_records:
                    pass
                else:
                    pass
                
                return trade_records
            
            # Helper function to extract NON-BANK LENDER CREDIT INFORMATION data
            def find_nlci_data():
                """Find data from NON-BANK LENDER CREDIT INFORMATION (NLCI) table only"""
                nlci_found = False
                nlci_start_idx = -1
                nlci_end_idx = -1
                result = {
                    'Ttl_Limit': "-",
                    'Ttl_Outstanding': "-",
                    'Conduct_Highest_Value': "-"
                }
                
                conduct_values = []  # Store all conduct values to find highest
                
                # First, find the exact boundaries of NLCI table
                for idx, row in df.iterrows():
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell).strip() if pd.notna(cell) else ""
                        # Must contain "NON-BANK LENDER CREDIT INFORMATION" and "NLCI" to be sure
                        if "NON-BANK LENDER CREDIT INFORMATION" in cell_str.upper() and "NLCI" in cell_str.upper():
                            nlci_found = True
                            nlci_start_idx = idx
                            break
                    
                    # Find where NLCI table ends (before WRITTEN-OFF ACCOUNT or next major section)
                    if nlci_found and nlci_end_idx == -1:
                        for cell in row:
                            cell_str = str(cell).strip() if pd.notna(cell) else ""
                            if "WRITTEN-OFF ACCOUNT" in cell_str.upper():
                                nlci_end_idx = idx
                                break
                    
                    if nlci_found and nlci_end_idx != -1:
                        break
                
                if not nlci_found:
                    return result
                
                # Set end boundary if not found
                if nlci_end_idx == -1:
                    nlci_end_idx = min(nlci_start_idx + 50, len(df))
                
                # Now extract data only from within NLCI table boundaries
                for search_idx in range(nlci_start_idx + 1, nlci_end_idx):
                    if search_idx >= len(df):
                        break
                    search_row = df.iloc[search_idx]
                    
                    # Convert row to list for easier processing
                    row_data = [str(cell).strip() if pd.notna(cell) else "" for cell in search_row]
                    
                    # Check for "TOTAL" row to extract Ttl_Limit and Ttl_Outstanding
                    # The row structure is: TOTAL  1,198.00  [empty cells]  TOTAL  675.28
                    total_count = 0
                    for col_idx, cell_str in enumerate(row_data):
                        if cell_str.upper() == "TOTAL":
                            total_count += 1
                            
                            # First TOTAL - get Ttl_Limit (next numeric value)
                            if total_count == 1:
                                for offset in range(1, min(10, len(row_data) - col_idx)):
                                    val = row_data[col_idx + offset].replace(',', '').strip()
                                    if val and re.match(r'^\d+\.?\d*$', val):
                                        result['Ttl_Limit'] = row_data[col_idx + offset].strip()
                                        break
                            
                            # Second TOTAL - get Ttl_Outstanding (next numeric value)
                            elif total_count == 2:
                                for offset in range(1, min(10, len(row_data) - col_idx)):
                                    val = row_data[col_idx + offset].replace(',', '').strip()
                                    if val and re.match(r'^\d+\.?\d*$', val):
                                        result['Ttl_Outstanding'] = row_data[col_idx + offset].strip()
                                        break
                                break  # Found both totals, exit loop
                    
                    # Collect all numeric values from "Conduct of Account" columns within NLCI section
                    # Look for rows that contain BNPL (Buy Now Pay Later) to confirm it's NLCI data
                    is_nlci_row = any("BNPL" in str(cell).upper() for cell in search_row)
                    
                    if is_nlci_row or any("OUTSTANDING CREDIT" in str(cell).upper() for cell in search_row):
                        # These appear in the right side of the table (after many columns)
                        for col_idx, cell_str in enumerate(row_data):
                            # Skip first few columns (No, date, capacity, etc.)
                            if col_idx > 10:  # Conduct columns are typically after column 10
                                val = cell_str.replace(',', '').strip()
                                # Check if it's a small integer (conduct values are typically 0-9)
                                if val and re.match(r'^\d+$', val):
                                    num_val = int(val)
                                    if 0 <= num_val <= 99:  # Reasonable range for conduct values
                                        conduct_values.append(num_val)
                
                # Find highest conduct value
                if conduct_values:
                    highest = max(conduct_values)
                    result['Conduct_Highest_Value'] = str(highest)
                else:
                    pass
                
                return result

            # Extract base data from all tables (common to all rows)
            base_row_data['Subject_Name'] = find_particulars_value("Name Of Subject")
            base_row_data['IC_PP_No'] = find_particulars_value("IC / PP No") 
            base_row_data['New_IC_No'] = find_particulars_value("New IC No")
            base_row_data['Your_Ref_No'] = find_particulars_value("Your Ref. No")
            base_row_data['Nationality'] = find_particulars_value("Nationality")
            
            # Extract SUMMARY CREDIT REPORT data with custom field naming
            summary_credit_data = find_summary_credit_data()
            base_row_data['A_App_No_Application'] = summary_credit_data['A_App_No_Application']
            base_row_data['A_App_Ttl_Amnt'] = summary_credit_data['A_App_Ttl_Amnt']
            base_row_data['B_Pend_No_Application'] = summary_credit_data['B_Pend_No_Application']
            base_row_data['B_Pend_Ttl_Amnt'] = summary_credit_data['B_Pend_Ttl_Amnt']
            
            # Extract SUMMARY OF POTENTIAL & CURRENT LIABILITIES data with hybrid structure
            potential_liabilities_data = find_potential_liabilities_data()
            base_row_data['AsBorr_Outstanding_RM'] = potential_liabilities_data['AsBorr_Outstanding_RM']
            base_row_data['AsBorr_Total_Limit_RM'] = potential_liabilities_data['AsBorr_Total_Limit_RM']
            base_row_data['AsBorr_FEC_Limit_RM'] = potential_liabilities_data['AsBorr_FEC_Limit_RM']
            base_row_data['Legal_Action_Taken'] = potential_liabilities_data['Legal_Action_Taken']
            base_row_data['Special_Attention_Account'] = potential_liabilities_data['Special_Attention_Account']
            
            # Extract Legal Suits and Bankruptcy data (totals in headers)
            legal_suits_bankruptcy_data = find_legal_suits_bankruptcy_data()
            base_row_data['Legal_Suits_Defendant'] = legal_suits_bankruptcy_data['Legal_Suits_Defendant']
            base_row_data['Legal_Suits_Plaintiff'] = legal_suits_bankruptcy_data['Legal_Suits_Plaintiff']
            base_row_data['Bankruptcy_Action'] = legal_suits_bankruptcy_data['Bankruptcy_Action']
            
            # Extract NON-BANK LENDER CREDIT INFORMATION data
            nlci_data = find_nlci_data()
            base_row_data['Ttl_Limit'] = nlci_data['Ttl_Limit']
            base_row_data['Ttl_Outstanding'] = nlci_data['Ttl_Outstanding']
            base_row_data['Conduct_Highest_Value'] = nlci_data['Conduct_Highest_Value']
            
            # Extract KEY STATISTICS data
            key_stats_data = find_key_statistics_data()
            base_row_data['SF_No_of_Facilities'] = key_stats_data['SF_No_of_Facilities']
            base_row_data['SF_Total_Outstanding_Balance_RM'] = key_stats_data['SF_Total_Outstanding_Balance_RM']
            base_row_data['SF_Total_Outstanding_Balance_Against_Total_Limit'] = key_stats_data['SF_Total_Outstanding_Balance_Against_Total_Limit']
            base_row_data['SF_Highest_No_of_Installments_Arrears_Last_12_months'] = key_stats_data['SF_Highest_No_of_Installments_Arrears_Last_12_months']
            base_row_data['UF_No_of_Facilities'] = key_stats_data['UF_No_of_Facilities']
            base_row_data['UF_Total_Outstanding_Balance_RM'] = key_stats_data['UF_Total_Outstanding_Balance_RM']
            base_row_data['UF_Total_Outstanding_Balance_Against_Total_Limit'] = key_stats_data['UF_Total_Outstanding_Balance_Against_Total_Limit']
            base_row_data['UF_Highest_No_of_Installments_Arrears_Last_12_months'] = key_stats_data['UF_Highest_No_of_Installments_Arrears_Last_12_months']
            base_row_data['CC_Average_Utilisation_Last_6_months'] = key_stats_data['CC_Average_Utilisation_Last_6_months']
            base_row_data['ORC_Average_Utilisation_Last_6_months'] = key_stats_data['ORC_Average_Utilisation_Last_6_months']
            base_row_data['CHC_Min_Utilisation_Last_12_months_RM'] = key_stats_data['CHC_Min_Utilisation_Last_12_months_RM']
            base_row_data['CHC_Max_Utilisation_Last_12_months_RM'] = key_stats_data['CHC_Max_Utilisation_Last_12_months_RM']
            base_row_data['NHEF_No_of_Accounts'] = key_stats_data['NHEF_No_of_Accounts']
            base_row_data['LL_No_of_Accounts'] = key_stats_data['LL_No_of_Accounts']
            base_row_data['FL_No_of_Accounts'] = key_stats_data['FL_No_of_Accounts']
            
            # Extract TRADE / CREDIT REFERENCE data (can be multiple records)
            trade_credit_records = find_trade_credit_reference_data()
            
            # Keep the remaining SUMMARY CREDIT INFORMATION fields (if they still exist in the PDF)
            base_row_data['Legal_Action_Banking'] = find_summary_credit_value("Legal Action taken (from Banking)")
            base_row_data['Existing_Facilities'] = find_summary_credit_value("Existing No. of Facility (from Banking)")
            base_row_data['Bankruptcy_Record'] = find_summary_credit_value("Bankruptcy Record")
            base_row_data['Legal_Suits'] = find_summary_credit_value("Legal Suits")
            base_row_data['Trade_Credit_Reference'] = find_summary_credit_value("Trade / Credit Reference")
            base_row_data['Total_Enquiries_12m'] = find_summary_credit_value("Total Enquiries for Last 12 months")
            base_row_data['Total_Companies_Interest'] = find_summary_credit_value("Total Companies/Businesses Interest")
            
            # Extract i-SCORE and calculate Risk Grade
            iscore_str = find_credit_score_value("i-SCORE")
            
            # Clean and extract numeric value from i-SCORE
            iscore_clean = ""
            iscore_num = 0
            if iscore_str:
                # Remove any non-numeric characters and extract numbers
                numbers = re.findall(r'\d+', str(iscore_str))
                if numbers:
                    iscore_clean = numbers[0]  # Take first number found
                    try:
                        iscore_num = int(iscore_clean)
                    except:
                        iscore_num = 0
                else:
                    pass
            
            # Calculate risk grade
            risk_grade = ""
            if iscore_num > 0:
                risk_grade_num = self.extract_risk_grade_from_score(iscore_num)
                risk_grade = str(risk_grade_num) if risk_grade_num else ""
            else:
                pass
            
            base_row_data['i_SCORE'] = iscore_clean if iscore_clean else iscore_str
            base_row_data['Risk_Grade'] = risk_grade
            
            # Extract CCRIS ENTITY data
            base_row_data['CCRIS_Entity_Key'] = find_ccris_entity_key()
            
            # Extract Subject Status data
            base_row_data['Warning_Remark'] = find_warning_remark()
            
            # Extract Key Contributing Factors
            contributing_factors_raw = find_credit_score_value("Key Contributing Factors")
            
            # Process contributing factors
            contributing_factors = []
            if contributing_factors_raw and len(contributing_factors_raw.strip()) > 0:
                clean_text = contributing_factors_raw.strip()
                
                if "•" in clean_text:
                    # Split by bullet points
                    parts = clean_text.split("•")
                    for part in parts:
                        cleaned_part = part.strip()
                        if cleaned_part and len(cleaned_part) > 3:
                            cleaned_part = re.sub(r'^[\s\-\*•]+', '', cleaned_part)
                            cleaned_part = re.sub(r'[\s\-\*•]+$', '', cleaned_part)
                            if cleaned_part:
                                contributing_factors.append(cleaned_part)
                else:
                    # Single factor
                    contributing_factors.append(clean_text)
            
            if not contributing_factors:
                contributing_factors.append("")  # Empty factor to maintain structure
            
            
            # Extract SHAREHOLDING INTEREST data
            shareholding_interests = find_shareholding_interests()
            
            # Extract KEY STATISTICS Earliest/Latest Approved Facilities
            earliest_facility, latest_facilities = find_key_statistics_facilities()
            
            # Determine row count: Use MAXIMUM of three tables
            # 1. Shareholding Interests (can be multiple)
            # 2. Trade/Credit Reference (can be multiple)
            # 3. Latest Facilities (up to 3)
            shareholding_count = len(shareholding_interests)
            trade_credit_count = len(trade_credit_records)
            latest_facilities_count = len(latest_facilities)
            
            max_row_count = max(shareholding_count, trade_credit_count, latest_facilities_count)
            
            
            # Combine contributing factors into single string (not multiplying)
            factors_combined = " | ".join(contributing_factors) if contributing_factors else "-"
            
            # Create rows based on maximum count
            result_rows = []
            
            for row_idx in range(max_row_count):
                row_data = base_row_data.copy()
                
                # Add contributing factor data (same for all rows)
                row_data['Key_Contributing_Factor'] = factors_combined
                
                # Add shareholding interest data (cycle through if not enough)
                if shareholding_count > 0:
                    interest = shareholding_interests[row_idx % shareholding_count]
                else:
                    interest = {'No': "-", 'Name': "-", 'Position': "-", 'Appointed': "-",
                               'Business_Expiry_Date': "-", 'Shareholding': "-", 'Percentage': "-",
                               'Remark': "-", 'Last_Updated_by_Experian': "-"}
                
                row_data['No'] = interest['No']
                row_data['Name'] = interest['Name']
                row_data['Position'] = interest['Position']
                row_data['Appointed'] = interest['Appointed']
                row_data['Business_Expiry_Date'] = interest['Business_Expiry_Date']
                row_data['Shareholding'] = interest['Shareholding']
                row_data['Percentage'] = interest['Percentage']
                row_data['Remark'] = interest['Remark']
                row_data['Last_Updated_by_Experian'] = interest['Last_Updated_by_Experian']
                
                # Add Trade/Credit Reference data (cycle through if not enough)
                if trade_credit_count > 0:
                    tcr = trade_credit_records[row_idx % trade_credit_count]
                else:
                    tcr = {'TCR_Subject_Name': "-", 'TCR_Creditors_Name': "-", 'TCR_Creditors_Contact': "-",
                          'TCR_Ref_No': "-", 'TCR_Industry': "-", 'TCR_Solicitors_Name': "-",
                          'TCR_Guarantor_Owner': "-", 'TCR_Subject_ID': "-", 'TCR_Amount_Due': "-",
                          'TCR_Aging_Days': "-", 'TCR_Debt_Type': "-", 'TCR_Document_Status_Date': "-",
                          'TCR_Solicitors_Contact': "-", 'TCR_Remark': "-"}
                
                row_data['TCR_Subject_Name'] = tcr['TCR_Subject_Name']
                row_data['TCR_Creditors_Name'] = tcr['TCR_Creditors_Name']
                row_data['TCR_Creditors_Contact'] = tcr['TCR_Creditors_Contact']
                row_data['TCR_Ref_No'] = tcr['TCR_Ref_No']
                row_data['TCR_Industry'] = tcr['TCR_Industry']
                row_data['TCR_Solicitors_Name'] = tcr['TCR_Solicitors_Name']
                row_data['TCR_Guarantor_Owner'] = tcr['TCR_Guarantor_Owner']
                row_data['TCR_Subject_ID'] = tcr['TCR_Subject_ID']
                row_data['TCR_Amount_Due'] = tcr['TCR_Amount_Due']
                row_data['TCR_Aging_Days'] = tcr['TCR_Aging_Days']
                row_data['TCR_Debt_Type'] = tcr['TCR_Debt_Type']
                row_data['TCR_Document_Status_Date'] = tcr['TCR_Document_Status_Date']
                row_data['TCR_Solicitors_Contact'] = tcr['TCR_Solicitors_Contact']
                row_data['TCR_Remark'] = tcr['TCR_Remark']
                
                # Add KEY STATISTICS Earliest Approved Facility (same for all rows)
                row_data['EAF_Facility_Type'] = earliest_facility['type']
                row_data['EAF_Date_Approved'] = earliest_facility['date']
                
                # Add KEY STATISTICS Latest Approved Facility (cycle through if not enough)
                if latest_facilities_count > 0:
                    latest_facility = latest_facilities[row_idx % latest_facilities_count]
                else:
                    latest_facility = {'type': "-", 'date': "-"}
                
                row_data['LAF_Facility_Type'] = latest_facility['type']
                row_data['LAF_Date_Approved'] = latest_facility['date']
                
                result_rows.append(row_data)
                
            
            return result_rows
            
        except Exception as e:
            print(f"Error extracting data from {file_name}: {e}")
            return None
    
    def update_database_preview(self, max_rows=200):
        """Update the database preview in the Database Viewer tab with improved layout"""
        try:
            # Clear existing columns and rows
            for col in self.db_tree["columns"]:
                self.db_tree.heading(col, text="")
            self.db_tree.delete(*self.db_tree.get_children())
            
            if self.database_df is None or self.database_df.empty:
                return
            
            # Configure columns
            columns = list(self.database_df.columns)
            self.db_tree['columns'] = columns
            
            # Configure column headings and equal widths for better layout
            standard_width = 150  # Equal width for all columns
            for col in columns:
                # Clean column name for display (remove underscores, title case)
                display_name = col.replace('_', ' ').title()
                self.db_tree.heading(col, text=display_name, anchor='w')
                self.db_tree.column(col, width=standard_width, minwidth=100, anchor='w')
            
            # Add data (limit to max_rows for performance)
            display_df = self.database_df.head(max_rows) if len(self.database_df) > max_rows else self.database_df
            
            for idx, row in display_df.iterrows():
                # Format values - replace empty/NaN with "-" for consistency
                values = []
                for col in columns:
                    cell_value = row[col]
                    if pd.isna(cell_value) or cell_value == "" or str(cell_value).strip() == "":
                        values.append("-")
                    else:
                        # Clean and format cell value
                        formatted_value = str(cell_value).strip()
                        # Limit very long values to prevent display issues
                        if len(formatted_value) > 50:
                            formatted_value = formatted_value[:47] + "..."
                        values.append(formatted_value)
                
                self.db_tree.insert("", "end", values=values)
        except Exception as e:
            print(f"update_database_preview error: {e}")
    
    def export_database(self):
        """Export the built database to an Excel file named Database_experian_<datetime>.xlsx"""
        try:
            if self.database_df is None or self.database_df.empty:
                messagebox.showinfo("Info", "No database to export. Please Build Database first.")
                return
            
            folder = filedialog.askdirectory(title="Select folder to save Database file")
            if not folder:
                return
            
            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"Database_experian_{now}.xlsx"
            out_path = Path(folder) / file_name
            
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                self.database_df.to_excel(writer, sheet_name="Database", index=False)
            
            messagebox.showinfo("Exported", f"Database exported to:\n{out_path}")
            self.status_label.configure(text=f"✓ Database exported: {out_path.name}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export database:\n{str(e)}")
            import traceback
            traceback.print_exc()

def main():
    app = PDFtoExcelApp()
    app.mainloop()

if __name__ == "__main__":
    main()

